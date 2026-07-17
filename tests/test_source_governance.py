from pathlib import Path

from openpyxl import Workbook

from strategyos_mvp.source_governance import (
    CONTROL_PLANE,
    CURRENT_EVIDENCE,
    EVALUATOR_ONLY,
    HISTORIC_CONTEXT,
    QUARANTINED_CONTEXT,
    RESTRICTED_CONTEXT,
    final_source_disposition,
    initial_source_disposition,
    is_agent_evidence_path,
    is_detector_candidate_path,
)
from strategyos_mvp.source_quality import inspect_workbook_formula_quality


def test_poc_control_and_evaluator_paths_override_content_classification():
    assert initial_source_disposition("02_Agent_JDs/Finance_Analyst_JD.md") == CONTROL_PLANE
    assert initial_source_disposition("03_Sample_Tasks/POC_Task_Brief.md") == CONTROL_PLANE
    assert initial_source_disposition("01_Synthetic_Dataset/README.md") == EVALUATOR_ONLY
    assert (
        initial_source_disposition(
            "01_Synthetic_Dataset/14_CEO_Office/CEO_Calendar_Mizan_Apr-Jul_2026.xlsx"
        )
        == RESTRICTED_CONTEXT
    )


def test_historic_paths_are_context_and_generic_ambiguity_is_quarantined():
    assert (
        initial_source_disposition(
            "01_Synthetic_Dataset/13_Historic_Correspondence/Email_4.txt"
        )
        == HISTORIC_CONTEXT
    )
    item = {
        "relative_path": "incoming/ambiguous.txt",
        "supported": True,
        "source_disposition": CURRENT_EVIDENCE,
        "classification": {"status": "ambiguous", "role": None},
    }
    assert final_source_disposition(item) == QUARANTINED_CONTEXT


def test_restricted_and_quarantined_context_never_enter_agent_or_detector_evidence():
    assert not is_agent_evidence_path("98_Restricted_Context/calendar.xlsx")
    assert not is_agent_evidence_path("97_Quarantined_Context/ambiguous.txt")
    assert is_agent_evidence_path("99_Historic_Context/history.csv")
    assert not is_detector_candidate_path("99_Historic_Context/history.csv")
    assert is_detector_candidate_path("02_ERP_Extracts/current.csv")


def test_formula_quality_detects_error_omission_and_local_pattern_break(tmp_path: Path):
    path = tmp_path / "forecast.xlsx"
    workbook = Workbook()
    forecast = workbook.active
    forecast.title = "Vendor_CF_Forecast"
    for column in range(3, 9):
        forecast.cell(row=4, column=column, value=column)
    forecast["I4"] = "=SUM(C4:G4)"
    forecast["D7"] = "=#REF!*0"
    cash = workbook.create_sheet("Cash_Position")
    for row in (104, 105, 106):
        cash.cell(row=row, column=5, value=2)
        cash.cell(row=row, column=6, value=3)
        cash.cell(row=row, column=7, value=f"=E{row}*F{row}")
    cash["G105"] = "=E102*F105"
    workbook.save(path)

    issues = inspect_workbook_formula_quality(path)
    by_code = {item["code"]: item for item in issues}

    assert by_code["formula_error_token"]["cell"] == "D7"
    assert by_code["sum_omits_adjacent_cell"]["cell"] == "I4"
    assert by_code["inconsistent_relative_formula"]["cell"] == "G105"
