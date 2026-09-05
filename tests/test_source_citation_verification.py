import hashlib
from types import SimpleNamespace
from openpyxl import Workbook
from strategyos_mvp.citation_resolver import verify_source_citations


def test_named_records_resolve_uniquely_and_changed_sources_fail(tmp_path):
    book=Workbook();sheet=book.active;sheet.title='Signals'
    sheet.append(['Signal_ID','Business Unit','Value']);sheet.append(['SIG-1','Company A',0]);sheet.append(['SIG-2','Company B',12])
    path=tmp_path/'signals.xlsx';book.save(path)
    evidence=SimpleNamespace(dataset_root=tmp_path,manifest={'signals.xlsx':{'sha256':hashlib.sha256(path.read_bytes()).hexdigest()}})
    bundle=SimpleNamespace(evidence=evidence,ap=None,ar=None,gl=None,trial_balance=None,vendors=None,customers=None,coa=None,po=None)
    rows,errors=verify_source_citations(bundle,[{'source_path':'signals.xlsx','locator':'SIG-1'}])
    assert not errors and rows[0]['locator']=='Signals!Excel row 2' and rows[0]['resolved']
    rows,errors=verify_source_citations(bundle,[{'source_path':'signals.xlsx','locator':'Business unit / Company B'}])
    assert not errors and rows[0]['locator']=='Signals!Excel row 3'
    assert not verify_source_citations(bundle,[{'source_path':'signals.xlsx','locator':'Signals!Excel rows 2-3'}])[1]
    assert verify_source_citations(bundle,[{'source_path':'signals.xlsx','locator':'Signals!Excel rows 2-4'}])[1]
    assert verify_source_citations(bundle,[{'source_path':'signals.xlsx','locator':'SIG-999'}])[1]
    assert verify_source_citations(bundle,[{'source_path':'../signals.xlsx','locator':'Signals!Excel row 2'}])[1]
    assert verify_source_citations(bundle,[{'source_path':'signals.xlsx','locator':'SIG-1','source_hash':'invented'}])[1]
    sheet.append(['SIG-1','Company A',99]);book.save(path)
    assert verify_source_citations(bundle,[{'source_path':'signals.xlsx','locator':'SIG-1'}])[1]
    evidence.manifest['signals.xlsx']['sha256']=hashlib.sha256(path.read_bytes()).hexdigest()
    assert verify_source_citations(bundle,[{'source_path':'signals.xlsx','locator':'SIG-1'}])[1]


def test_ambiguous_id_has_explicit_repair_candidates_without_silent_selection(tmp_path):
    import hashlib
    from types import SimpleNamespace
    from openpyxl import Workbook
    from strategyos_mvp.citation_resolver import canonical_workbook_locator, citation_location_hints
    book = Workbook(); book.active.title = 'Register'
    book.active.append(['KPI_ID', 'Month', 'Actual']); book.active.append(['KPI-09', '2026-05', 35]); book.active.append(['KPI-09', '2026-06', 36.2])
    path = tmp_path / 'kpis.xlsx'; book.save(path)
    bundle = SimpleNamespace(evidence=SimpleNamespace(dataset_root=tmp_path, manifest={'kpis.xlsx': {'sha256': hashlib.sha256(path.read_bytes()).hexdigest()}}))
    assert canonical_workbook_locator(bundle, 'kpis.xlsx', 'KPI-09') is None
    hints = citation_location_hints(bundle, [{'source_path': 'kpis.xlsx', 'locator': 'KPI-09'}])
    assert hints[0]['candidate_count'] == 2
    assert 'Register!Excel row 3' in hints[0]['candidates']
    assert '2026-06' in hints[0]['candidates']
    assert 'UNTRUSTED' in hints[0]['candidates']


def test_targeted_payroll_coverage_is_hash_checked_and_question_bounded(tmp_path):
    import hashlib
    from types import SimpleNamespace
    from openpyxl import Workbook
    from strategyos_mvp.source_search import targeted_financial_records
    path = tmp_path / 'Headcount_Payroll_2026.xlsx'
    book = Workbook(); book.active.append(['Quarter', 'Headcount']); book.active.append(['2026-Q1', 100]); book.save(path)
    evidence = SimpleNamespace(dataset_root=tmp_path, manifest={path.name: {'sha256': hashlib.sha256(path.read_bytes()).hexdigest()}})
    assert not targeted_financial_records(evidence, 'What is cash?')['records']
    result = targeted_financial_records(evidence, 'What is revenue per employee?')
    assert result['coverage_complete'] and result['records'][0]['source_path'] == path.name
    path.write_bytes(b'changed')
    assert targeted_financial_records(evidence, 'What is headcount?')['status'] == 'unavailable'


def test_targeted_synergy_charter_retains_every_role_in_source_order(tmp_path):
    import hashlib
    from types import SimpleNamespace
    from zipfile import ZipFile
    from strategyos_mvp.source_search import targeted_financial_records
    rows=[('Programme','Owner'),('SYN-01','Group COO'),('SYN-02','Group CIO'),('SYN-03','Group Property')]
    xml='<w:document xmlns:w="http://schemas.openxmlformats.org/wordprocessingml/2006/main"><w:body><w:tbl>'
    xml+=''.join('<w:tr>'+''.join('<w:tc><w:p><w:r><w:t>'+value+'</w:t></w:r></w:p></w:tc>' for value in row)+'</w:tr>' for row in rows)
    xml+='</w:tbl></w:body></w:document>'
    path=tmp_path/'Synergy_Programme_Charter.docx'
    with ZipFile(path,'w') as archive: archive.writestr('word/document.xml',xml)
    evidence=SimpleNamespace(dataset_root=tmp_path,manifest={path.name:{'sha256':hashlib.sha256(path.read_bytes()).hexdigest()}})
    result=targeted_financial_records(evidence,'Who owns synergy delivery?')
    text='\n'.join(row['text'] for row in result['records'])
    assert result['coverage_complete'] and text.index('Group COO') < text.index('Group CIO') < text.index('Group Property')
    assert all(row['locator'].startswith('Document paragraph ') for row in result['records'])
