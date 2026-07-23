from datetime import date, timedelta

from openpyxl import Workbook

from strategyos_mvp.source_calendar import derive_calendar_agenda


def test_calendar_projection_processes_all_rows_and_prioritises_upcoming(tmp_path):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Calendar"
    sheet.append(
        [
            "Date",
            "Day",
            "Start",
            "End",
            "Title",
            "Attendees",
            "Location",
            "Category",
            "Related BU",
            "Past/Upcoming",
            "Notes / Agenda",
        ]
    )
    current_day = date.today()
    event_days = [current_day - timedelta(days=2), current_day - timedelta(days=1)] + [
        current_day + timedelta(days=offset) for offset in range(1, 15)
    ]
    for index, event_day in enumerate(event_days, start=1):
        sheet.append(
            [
                event_day.isoformat(),
                event_day.strftime("%A"),
                "09:00",
                "10:00",
                f"Executive event {index}",
                "CEO; CFO",
                "Boardroom",
                "Operating cadence",
                "Group",
                "Upcoming" if event_day >= current_day else "Past",
                f"Review decision brief {index}.",
            ]
        )
    path = tmp_path / "CEO_Calendar.xlsx"
    workbook.save(path)

    agenda = derive_calendar_agenda(tmp_path)

    assert agenda["status"] == "ready"
    assert agenda["total_item_count"] == 16
    assert agenda["upcoming_item_count"] == 14
    assert agenda["projection_as_of"] == current_day.isoformat()
    assert agenda["projection_policy"] == "upcoming_first"
    assert len(agenda["items"]) == 12
    first = agenda["items"][0]
    assert first["date"] == (current_day + timedelta(days=1)).isoformat()
    assert first["when"] == "09:00"
    assert first["ends_at"] == "10:00"
    assert first["attendees"] == "CEO; CFO"
    assert first["location"] == "Boardroom"
    assert first["prep"] == "Review decision brief 3."


def test_calendar_projection_excludes_personal_and_ambiguous_entries(tmp_path):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Calendar"
    sheet.append(["Date", "Title", "Category", "Related BU", "Notes / Agenda"])
    tomorrow = date.today() + timedelta(days=1)
    sheet.append(
        [
            tomorrow.isoformat(),
            "Executive Committee",
            "Decision meeting",
            "Group",
            "Review the board decision brief.",
        ]
    )
    sheet.append(
        [
            tomorrow.isoformat(),
            "Anniversary dinner reservation",
            "Personal",
            "",
            "Book the florist.",
        ]
    )
    sheet.append(
        [
            tomorrow.isoformat(),
            "Coffee catch-up",
            "Social",
            "",
            "",
        ]
    )
    workbook.save(tmp_path / "CEO_Calendar.xlsx")

    agenda = derive_calendar_agenda(tmp_path)

    assert agenda["status"] == "ready"
    assert agenda["total_item_count"] == 1
    assert agenda["excluded_non_business_count"] == 2
    assert [item["title"] for item in agenda["items"]] == ["Executive Committee"]
    assert agenda["items"][0]["business_relevant"] is True
    assert agenda["items"][0]["relevance_reason"] == "deterministic_business_marker"


def test_calendar_projection_reports_when_all_complete_rows_are_excluded(tmp_path):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Calendar"
    sheet.append(["Date", "Title", "Category"])
    sheet.append([(date.today() + timedelta(days=1)).isoformat(), "Dentist appointment", "Personal"])
    workbook.save(tmp_path / "CEO_Calendar.xlsx")

    agenda = derive_calendar_agenda(tmp_path)

    assert agenda["status"] == "unavailable"
    assert agenda["items"] == []
    assert agenda["excluded_non_business_count"] == 1
    assert agenda["reason"] == "No calendar item has been classified as business-relevant for the CEO projection."
