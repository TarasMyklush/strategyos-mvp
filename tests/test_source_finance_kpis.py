from __future__ import annotations

from pathlib import Path

from strategyos_mvp.executive_presentation import build_executive_presentation
from strategyos_mvp.executive_read_model import build_executive_read_model
from strategyos_mvp.source_finance_kpis import derive_source_finance_kpis


DATASET = Path(__file__).parent / "fixtures" / "01_Synthetic_Dataset"


def test_source_pack_calculates_four_ceo_actuals_with_lineage():
    payload = derive_source_finance_kpis(DATASET)

    assert payload["authoritative"] is True
    assert payload["derived_from"] == "deterministic_source_finance_kpi_engine"
    assert payload["reporting_period_key"] == "H1 2026"
    assert payload["components"] == {
        "revenue_actual": "385079908.90",
        "revenue_plan": None,
        "cogs_actual": "75503688.29",
        "ebitda_actual": "215741310.56",
        "ebitda_plan": None,
        "operating_cost_actual": "93834910.05",
        "operating_cost_plan": None,
        "cash_balance": "42341408.58",
        "board_floor": None,
    }
    assert payload["evidence"]["revenue"]["details"]["account_scopes"]["revenue"]["accounts"] == [
        "4000", "4010", "4020", "4030"
    ]
    revenue_drivers = payload["evidence"]["revenue"]["details"]["contributors"]["revenue"]
    assert [(item["label"], item["value_sar"], item["share_pct"]) for item in revenue_drivers] == [
        ("Revenue – Catering", "123016434.85", 31.9),
        ("Revenue – Government", "109896978.70", 28.5),
        ("Revenue – Modern Trade", "103168943.62", 26.8),
        ("Revenue – Hospitality", "48997551.73", 12.7),
    ]
    assert payload["evidence"]["cash_vs_floor"]["actual_complete"] is False
    assert payload["evidence"]["cash_vs_floor"]["details"]["missing_accounts"] == ["Emirates NBD EUR"]
    revenue_trend = payload["trend"]["revenue"]
    assert revenue_trend["labels"] == ["2026-01", "2026-02", "2026-03", "2026-04", "2026-05", "2026-06"]
    assert revenue_trend["actual"] == [
        "72561833.68",
        "56571935.86",
        "67615475.08",
        "68261922.69",
        "56702067.56",
        "63366674.03",
    ]
    assert revenue_trend["plan"] == []
    assert revenue_trend["has_plan_series"] is False
    assert revenue_trend["unit"] == "sar"
    ebitda_margin_trend = payload["trend"]["ebitda_margin"]
    assert ebitda_margin_trend["labels"] == revenue_trend["labels"]
    assert ebitda_margin_trend["actual"] == ["44.06", "38.77", "55.40", "64.10", "60.95", "72.69"]
    assert ebitda_margin_trend["plan"] == []
    assert ebitda_margin_trend["has_plan_series"] is False
    assert ebitda_margin_trend["unit"] == "percent"
    assert payload["dynamics"]["revenue"]["lifting"][0]["delta"].startswith("+SAR ")
    assert "5615726.86" not in payload["dynamics"]["revenue"]["lifting"][0]["delta"]


def test_source_pack_actuals_render_without_inventing_missing_comparators():
    read_model = build_executive_read_model(
        {"run_id": "source-pack-run", "finance_kpi": derive_source_finance_kpis(DATASET)},
        [], {}, {"report_count": 0}, {},
    )
    cards = build_executive_presentation(read_model)["driver_grid"]

    assert [(card["label"], card["metric"]) for card in cards] == [
        ("Revenue", "SAR 385.1M"),
        ("EBITDA margin", "56.0%"),
        ("Operating cost", "SAR 93.8M"),
        ("Cash vs floor", "SAR 42.3M"),
    ]
    assert cards[0]["availability"] == "partial"
    assert cards[0]["trend"]["actual"]
    assert cards[0]["trend"]["plan"] == []
    assert cards[0]["trend"]["has_plan_series"] is False
    assert cards[0]["missing_inputs"] == ["H1 budget aligned to this reporting scope"]
    assert cards[1]["missing_inputs"] == [
        "H1 EBITDA budget aligned to this scope",
        "H1 revenue budget aligned to this scope",
    ]
    assert cards[3]["missing_inputs"] == [
        "Approved cash floor aligned to this reporting scope",
        "Complete latest cash-position balances",
    ]
    assert [card["executive_brief"]["strategic_reference"] for card in cards] == [None, None, None, None]
    assert [(card["ring_pct"], card["ring_label"]) for card in cards] == [
        (None, ""),
        (56.0, "current margin"),
        (24.4, "of revenue"),
        (None, ""),
    ]
    assert cards[1]["trend"]["unit"] == "percent"
    assert cards[1]["trend"]["actual"] == [44.06, 38.77, 55.4, 64.1, 60.95, 72.69]
    assert cards[0]["source_files"] == [
        "02_ERP_Extracts/GL_Extract_H1_2026.csv",
        "03_Master_Data/Chart_of_Accounts.xlsx",
    ]
    assert "8,479 GL rows" in cards[0]["evidence_summary"]
    assert "missing Emirates NBD EUR" in cards[3]["detail"]
    assert cards[1]["executive_brief"]["calculation"]["steps"] == [
        {"label": "Revenue", "value": "SAR 385.1M"},
        {"label": "Less cost of goods sold", "value": "SAR 75.5M"},
        {"label": "Less operating cost", "value": "SAR 93.8M"},
        {"label": "EBITDA", "value": "SAR 215.7M"},
    ]
    assert cards[2]["executive_brief"]["readout"] == (
        "The current actual is available, but a CEO performance conclusion is not yet safe."
    )
    assert cards[2]["executive_brief"]["executive_signal"]["posture"] == "Comparison pending"
    assert "Group CFO" in cards[2]["executive_brief"]["decision_context"]
    assert cards[2]["executive_brief"]["drivers"][0]["label"] == "Salaries & Wages"
    assert cards[2]["executive_brief"]["decision_context"].startswith("Keep this with the Group CFO")


def _write_reconciliation(path: Path, *, division_2026f: str = "775") -> None:
    """A minimal division-to-group reconciliation, matching the real file's shape."""
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "Division_to_Group"
    ws.append(["SAR M", "2023A", "2024A", "2025A", "2026F"])
    ws.append(["Central Region division net revenue (this ERP dataset)", 590, 633, 697, division_2026f])
    ws.append(["Tamween Pharma Distribution BU total", 1940, 2120, 2340, 2540])
    wb.save(path)


def test_reconciliation_plan_derives_the_aligned_h1_budget(tmp_path):
    """The dataset carries a division 2026F forecast; the H1 plan is half of it.

    This is the fix for the live "plan comparison unavailable" -- the budget was
    there all along in the reconciliation file, just never read.
    """
    from strategyos_mvp.source_finance_kpis import _reconciliation_plan

    _write_reconciliation(tmp_path / "Division_to_Group_Reconciliation.xlsx")
    plan = _reconciliation_plan(tmp_path, "H1 2026")
    assert plan is not None
    # 775M annual x 0.5 for H1
    assert plan["revenue_plan"] == "387500000.00"
    assert plan["basis"]["forecast_column"] == "2026F"
    assert plan["basis"]["annual_plan_sar"] == "775000000.00"
    assert plan["basis"]["period_fraction"] == "0.5"


def test_plan_is_absent_when_the_dataset_carries_no_reconciliation(tmp_path):
    """No file -> fail closed, exactly as the old dataset does."""
    from strategyos_mvp.source_finance_kpis import _reconciliation_plan

    assert _reconciliation_plan(tmp_path, "H1 2026") is None


def test_plan_is_absent_for_a_period_it_cannot_align(tmp_path):
    """An unknown period shape must not silently borrow the H1 halving."""
    from strategyos_mvp.source_finance_kpis import _reconciliation_plan

    _write_reconciliation(tmp_path / "Division_to_Group_Reconciliation.xlsx")
    assert _reconciliation_plan(tmp_path, "mystery period") is None
    # A different year has no matching forecast column.
    assert _reconciliation_plan(tmp_path, "H1 2029") is None


def test_a_full_year_period_takes_the_whole_plan_not_half(tmp_path):
    from strategyos_mvp.source_finance_kpis import _reconciliation_plan

    _write_reconciliation(tmp_path / "Division_to_Group_Reconciliation.xlsx")
    plan = _reconciliation_plan(tmp_path, "FY 2026")
    assert plan["revenue_plan"] == "775000000.00"
    assert plan["basis"]["period_fraction"] == "1"


def test_existing_dataset_without_a_plan_still_reports_none():
    """The shared fixture has no reconciliation file: plan must stay None."""
    payload = derive_source_finance_kpis(DATASET)
    assert payload["components"]["revenue_plan"] is None
    assert payload["trend"]["revenue"]["has_plan_series"] is False


def test_group_budget_becomes_the_ceo_projection_without_double_counting_total(tmp_path):
    """The calculated GROUP row is the headline, never an additional BU."""
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "BU_Budget_2026"
    ws.append([
        "Business Unit", "H1 Budget", "H1 Actual/Est (SAR M)", "H1 Var",
        "EBITDA Budget %", "EBITDA H1 Est %", "Note",
    ])
    ws.append(["North", 100, 105, 5, 10, 12, "Demand ahead"])
    ws.append(["South", 200, 195, -5, 20, 18, "Mix pressure"])
    # This is already the consolidated figure: a regression must not sum it
    # with North and South, which would report 800M instead of 300M.
    ws.append(["GROUP", 300, 300, 0, 16.7, 16, "Consolidated"])
    cash = wb.create_sheet("Group_Cash_Floor")
    cash.append(["Quarter", "Group cash budget (SAR B)", "Actual/Forecast (SAR B)", "Floor (SAR B)"])
    cash.append(["2026-Q2 (est)", 1.2, 1.4, 1.0])
    wb.save(tmp_path / "BU_Group_Budget_2026.xlsx")

    payload = derive_source_finance_kpis(tmp_path)

    assert payload["components"] == {
        "revenue_actual": "300000000.00",
        "revenue_plan": "300000000.00",
        "ebitda_actual": "48000000.00",
        "ebitda_plan": "50100000.00",
        "operating_cost_actual": "252000000.00",
        "operating_cost_plan": "249900000.00",
        "cash_balance": "1400000000.00",
        "board_floor": "1000000000.00",
    }
    assert [item["name"] for item in payload["dynamics"]["revenue"]["lifting"]] == ["North"]
    assert [item["name"] for item in payload["dynamics"]["revenue"]["dragging"]] == ["South"]
    assert [item["name"] for item in payload["dynamics"]["ebitda_margin"]["lifting"]] == ["North"]
    assert [item["name"] for item in payload["dynamics"]["ebitda_margin"]["dragging"]] == ["South"]
    assert payload["dynamics"]["operating_cost"]["lifting"]
    assert payload["dynamics"]["operating_cost"]["dragging"]


def test_group_cash_floor_history_is_a_governed_chart_not_a_synthetic_series():
    from openpyxl import Workbook

    from strategyos_mvp.source_finance_kpis import _group_cash_floor_trend

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Group_Cash_Floor"
    sheet.append([
        "Quarter",
        "Group cash budget (SAR B)",
        "Actual/Forecast (SAR B)",
        "Floor (SAR B)",
        "Headroom",
        "Note",
    ])
    sheet.append(["2025-Q4 (actual)", 1.30, 1.32, 1.20, 0.12, "Approved floor"])
    sheet.append(["2026-Q1 (actual)", 1.34, 1.37, 1.20, 0.17, "Approved floor"])
    sheet.append(["2026-Q2 (est)", 1.38, 1.41, 1.20, 0.21, "Approved floor"])

    trend = _group_cash_floor_trend(workbook)

    assert trend == {
        "labels": ["2025-Q4", "2026-Q1", "2026-Q2"],
        "actual": ["1320000000.00", "1370000000.00", "1410000000.00"],
        "plan": ["1300000000.00", "1340000000.00", "1380000000.00"],
        "has_plan_series": True,
        "unit": "sar",
        "scope_note": "Quarterly group cash actual/forecast versus budget; the approved floor remains the headline comparator.",
        "plan_note": "Approved quarterly group cash budget from Group_Cash_Floor.",
    }


def test_division_monthly_budget_aligns_plan_to_actual_account_scope(tmp_path):
    from openpyxl import Workbook

    from strategyos_mvp.source_finance_kpis import _division_monthly_plan

    workbook = Workbook()
    revenue = workbook.active
    revenue.title = "Revenue_Budget_Monthly"
    revenue.append(["Line", "2026-01", "2026-02"])
    revenue.append(["Revenue — Modern Pharmacy (4000)", 100, 120])
    revenue.append(["Unrelated revenue (4099)", 900, 900])
    cost = workbook.create_sheet("Cost_Budget_Monthly")
    cost.append(["Line", "2026-01", "2026-02"])
    cost.append(["COGS (5000)", 30, 40])
    cost.append(["Operating cost (6000)", 20, 20])
    cost.append(["Depreciation outside EBITDA scope (6500)", 80, 80])
    path = tmp_path / "Division_Budget_2026.xlsx"
    workbook.save(path)

    plan = _division_monthly_plan(
        path,
        labels=["2026-01", "2026-02"],
        revenue_accounts=["4000"],
        cogs_accounts=["5000"],
        operating_cost_accounts=["6000"],
    )

    assert plan == {
        "revenue": ["100.00", "120.00"],
        "ebitda_margin": ["50.00", "50.00"],
        "operating_cost": ["50.00", "60.00"],
    }


def test_group_cash_movers_are_populated_from_governed_floor_headroom():
    from openpyxl import Workbook

    from strategyos_mvp.source_finance_kpis import _group_cash_floor_movers

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Group_Cash_Floor"
    sheet.append(["Quarter", "Actual/Forecast (SAR B)", "Floor (SAR B)", "Note"])
    sheet.append(["2026-Q1 (actual)", 1.37, 1.20, "Approved floor"])
    sheet.append(["2026-Q2 (est)", 1.41, 1.20, "Approved floor"])

    movers = _group_cash_floor_movers(workbook)

    assert movers["dragging"] == []
    assert [(item["name"], item["delta"]) for item in movers["lifting"]] == [
        ("2026-Q2 group cash", "SAR 210.0M above floor"),
        ("2026-Q1 group cash", "SAR 170.0M above floor"),
    ]
