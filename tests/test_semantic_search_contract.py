import hashlib
import json
from types import SimpleNamespace
import pytest
from openpyxl import Workbook
from strategyos_mvp import semantic_embeddings, source_search, vector_store


def test_source_rows_preserve_zero_exact_location_and_detect_file_mutation(tmp_path):
    book=Workbook();book.active.title='Budget';book.active.append(['BU','Actual']);book.active.append(['A',0])
    path=tmp_path/'budget.xlsx';book.save(path)
    evidence=SimpleNamespace(dataset_root=tmp_path,manifest={'budget.xlsx':{'sha256':hashlib.sha256(path.read_bytes()).hexdigest()}},pdf_text={})
    rows=list(source_search.source_records(evidence))
    assert rows[0][2:] == ('Budget!Excel row 2','BU: A; Actual: 0')
    book.active['B2']=999;book.save(path)
    with pytest.raises(ValueError,match='changed'):
        list(source_search.source_records(evidence))


def test_dense_search_sends_scope_filter_and_query_embedding(monkeypatch):
    calls=[]
    monkeypatch.setattr(semantic_embeddings,'configured',lambda:True)
    monkeypatch.setattr(vector_store,'_ensure_collection',lambda *args:None)
    monkeypatch.setattr(vector_store,'_embed_text',lambda text,query: [1.0] if query else (_ for _ in ()).throw(AssertionError()))
    monkeypatch.setattr(vector_store,'_qdrant_request',lambda method,path,payload: calls.append((method,path,payload)) or {'result':[]})
    scope={'must':[{'key':'tenant_slug','match':{'value':'tenant-a'}},{'key':'run_id','match':{'value':'run-a'}}]}
    vector_store._search_collection(vector_store.COLLECTION_NAME,query='المورد',filter_payload=scope,limit=3,create=True)
    assert calls[0][2]['filter']==scope
    assert calls[0][2]['vector']==[1.0]
    assert calls[0][1].endswith('/points/search')


def test_configured_semantic_status_never_substitutes_legacy_index(monkeypatch):
    monkeypatch.setattr(vector_store,'CONFIG',SimpleNamespace(qdrant_url='http://example'))
    monkeypatch.setattr(semantic_embeddings,'configured',lambda:True)
    calls=[]
    monkeypatch.setattr(vector_store,'_vector_status_for_collection',lambda run,collection,create: calls.append(collection) or {'status':'empty'})
    assert vector_store.vector_status_for_run('run-a')['status']=='empty'
    assert calls==[vector_store.COLLECTION_NAME]


def test_missing_model_manifest_fails_without_runtime_download(tmp_path,monkeypatch):
    monkeypatch.setenv('STRATEGYOS_EMBEDDING_MODEL_PATH',str(tmp_path))
    semantic_embeddings.model.cache_clear()
    with pytest.raises(FileNotFoundError):semantic_embeddings.model()
    (tmp_path/'model-manifest.json').write_text(json.dumps({'model_name':'wrong','revision':'wrong'}))
    with pytest.raises(RuntimeError,match='identity'):semantic_embeddings.model()
    semantic_embeddings.model.cache_clear()


def test_write_only_workbook_without_dimension_metadata_is_indexed(tmp_path):
    book=Workbook(write_only=True);sheet=book.create_sheet('Measurements');sheet.append(['Metric','Value']);sheet.append(['Zero',0]);path=tmp_path/'stream.xlsx';book.save(path)
    evidence=SimpleNamespace(dataset_root=tmp_path,manifest={'stream.xlsx':{'sha256':hashlib.sha256(path.read_bytes()).hexdigest()}},pdf_text={})
    rows=list(source_search.source_records(evidence))
    assert rows[0][2:] == ('Measurements!Excel row 2','Metric: Zero; Value: 0')


def test_reindex_reuses_only_matching_scope_and_hash(monkeypatch):
    monkeypatch.setattr(semantic_embeddings,'configured',lambda:True)
    monkeypatch.setattr(source_search,'source_records',lambda evidence:iter([('a.xlsx','hash','Sheet!Excel row 2','A: 0')]))
    monkeypatch.setattr(vector_store,'_run_filter',lambda run:None)
    monkeypatch.setattr(vector_store,'_ensure_collection',lambda:None)
    embedded=[];writes=[]
    monkeypatch.setattr(semantic_embeddings,'embed_many',lambda texts:embedded.extend(texts) or [[1.0] for text in texts])
    def request(method,path,payload):
        if path.endswith('/points'):
            return {'result':[{'id':payload['ids'][0],'payload':{'run_id':'run','tenant_slug':'tenant','source_hash':'hash'}}]}
        if method == 'PUT': writes.append(payload)
        return {}
    monkeypatch.setattr(vector_store,'_qdrant_request',request)
    result=source_search.sync_sources(run_id='run',tenant_slug='tenant',evidence=object())
    assert result['reused_points']==1 and not embedded and not writes
    result=source_search.sync_sources(run_id='run',tenant_slug='other',evidence=object())
    assert result['reused_points']==0 and len(embedded)==1 and len(writes)==1


def test_source_index_includes_text_briefings_and_office_paragraphs(tmp_path):
    from zipfile import ZipFile
    (tmp_path/'brief.txt').write_text('Inventory build requires working capital. اختبار')
    with ZipFile(tmp_path/'charter.docx','w') as archive:
        archive.writestr('word/document.xml','<w:document xmlns:w="http://schemas.openxmlformats.org/wordprocessingml/2006/main"><w:body><w:p><w:r><w:t>Owner: review committee</w:t></w:r></w:p></w:body></w:document>')
    manifest={path.name:{'sha256':hashlib.sha256(path.read_bytes()).hexdigest()} for path in tmp_path.iterdir()}
    evidence=SimpleNamespace(dataset_root=tmp_path,manifest=manifest,pdf_text={})
    rows=list(source_search.source_records(evidence))
    assert {row[0] for row in rows}=={'brief.txt','charter.docx'}
    assert rows[0][2]=='Text chunk 1' and 'اختبار' in rows[0][3]
    assert rows[1][2:] == ('Document paragraph 1','Owner: review committee')


def test_source_citations_resolve_streamed_rows_and_exact_text_locations(tmp_path):
    from strategyos_mvp.citation_resolver import resolve_citation
    from strategyos_mvp.models import Citation
    book=Workbook(write_only=True);sheet=book.create_sheet('Hedges');sheet.append(['Value']);sheet.append([0])
    relative='07_Cash_Forecast/CFO_Cash_Forecast_June_2026.xlsx';path=tmp_path/relative;path.parent.mkdir();book.save(path)
    (tmp_path/'brief.txt').write_text('Verified brief with zero change.')
    evidence=SimpleNamespace(dataset_root=tmp_path,manifest={p:{'sha256':hashlib.sha256((tmp_path/p).read_bytes()).hexdigest()} for p in (relative,'brief.txt')},pdf_text={})
    bundle=SimpleNamespace(evidence=evidence)
    for source,locator in ((relative,'Hedges!Excel row 2'),('brief.txt','Text chunk 1')):
        assert resolve_citation(bundle,Citation(source,locator,source_hash=evidence.manifest[source]['sha256']))['resolved']
    assert not resolve_citation(bundle,Citation('brief.txt','Text chunk 999',source_hash=evidence.manifest['brief.txt']['sha256']))['resolved']


def test_question_bank_is_not_indexed_even_with_an_arbitrary_filename(tmp_path):
    book=Workbook();book.active.append(['Question','Answer type','Where StrategyOS finds the answer']);book.active.append(['What is revenue?','Lookup','Budget'])
    path=tmp_path/'business.xlsx';book.save(path)
    evidence=SimpleNamespace(dataset_root=tmp_path,manifest={path.name:{'sha256':hashlib.sha256(path.read_bytes()).hexdigest()}},pdf_text={})
    assert list(source_search.source_records(evidence))==[]
