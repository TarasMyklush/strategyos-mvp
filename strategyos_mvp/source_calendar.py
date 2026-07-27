"""Governed CEO-agenda extraction from an optional uploaded calendar workbook."""

from __future__ import annotations

from datetime import date, datetime, timedelta
from pathlib import Path
import re
from typing import Any

from openpyxl import load_workbook

from .source_governance import RESTRICTED_CONTEXT_DIR


_PERSONAL_TERMS = frozenset(
    {
        "anniversary",
        "birthday",
        "dentist",
        "doctor",
        "family",
        "florist",
        "holiday",
        "medical",
        "personal",
        "private",
        "school",
        "vacation",
        "wedding",
    }
)
_BUSINESS_TERMS = frozenset(
    {
        "board",
        "budget",
        "business",
        "client",
        "committee",
        "compliance",
        "contract",
        "decision",
        "executive",
        "finance",
        "governance",
        "investor",
        "leadership",
        "meeting",
        "operating",
        "review",
        "strategy",
        "vendor",
    }
)


def _workbook_projection_anchor(workbook: Any) -> date | None:
    """Read an explicit governed 'today' anchor from workbook metadata.

    Synthetic and point-in-time source packs may intentionally model a date
    other than the application server's wall clock.  The workbook remains the
    authority when it declares that anchor; ordinary live calendars without
    one continue to use ``date.today()``.
    """

    anchor_pattern = re.compile(
        r"\btoday['’]?\s+anchor\s*:\s*(?:[A-Za-z]+[,\s]+)?"
        r"(?P<date>\d{4}-\d{2}-\d{2}|\d{1,2}\s+[A-Za-z]+\s+\d{4})",
        re.IGNORECASE,
    )
    for sheet in workbook.worksheets:
        for row in sheet.iter_rows(values_only=True):
            for raw in row:
                if raw is None:
                    continue
                match = anchor_pattern.search(str(raw))
                if not match:
                    continue
                value = match.group("date")
                for pattern in ("%Y-%m-%d", "%d %B %Y", "%d %b %Y"):
                    try:
                        return datetime.strptime(value, pattern).date()
                    except ValueError:
                        continue
    return None


def _business_relevance(
    *,
    explicit_value: Any,
    title: str,
    event_type: str,
    related_bu: str,
) -> tuple[bool, str]:
    """Return a deterministic, privacy-first calendar relevance decision.

    We never use a language model at source ingestion time. An explicit source
    classification wins. Otherwise a personal marker is excluded; only a
    positively identifiable business marker is included. Ambiguous events stay
    out of the CEO agenda until their source is classified, which avoids
    surfacing personal calendar data by inference.
    """

    normalized_explicit = str(explicit_value or "").strip().casefold()
    if normalized_explicit in {"1", "true", "yes", "y", "business", "include", "included"}:
        return True, "source_classified_business"
    if normalized_explicit in {"0", "false", "no", "n", "personal", "private", "exclude", "excluded"}:
        return False, "source_classified_non_business"

    text = " ".join((title, event_type, related_bu)).casefold()
    if any(term in text for term in _PERSONAL_TERMS):
        return False, "deterministic_personal_marker"
    if any(term in text for term in _BUSINESS_TERMS):
        return True, "deterministic_business_marker"
    return False, "requires_source_classification"


def derive_calendar_agenda(dataset_root: Path) -> dict[str, Any]:
    root = Path(dataset_root)
    candidates = sorted(path for path in root.rglob("*.xlsx") if "calendar" in path.name.lower() or "agenda" in path.name.lower())
    if not candidates:
        return {"status": "unavailable", "items": [], "reason": "No governed calendar workbook was supplied for this run."}
    path = candidates[0]
    relative_source = path.relative_to(root).as_posix()
    restricted = RESTRICTED_CONTEXT_DIR in path.relative_to(root).parts
    workbook = load_workbook(path, data_only=True, read_only=True)
    workbook_anchor = _workbook_projection_anchor(workbook)
    sheet = next((item for item in workbook.worksheets if item.title.strip().lower() in {"calendar", "agenda"}), workbook.active)
    rows = iter(sheet.values)
    headers = {_header_key(value): index for index, value in enumerate(next(rows, ())) if value is not None}
    def value(values: tuple[Any, ...], *names: str) -> Any:
        index = next((headers[name] for name in names if name in headers), None)
        return values[index] if index is not None and index < len(values) else None
    items: list[dict[str, Any]] = []
    excluded_count = 0
    for raw in rows:
        values = tuple(raw)
        event_date = _date(value(values, "event_date", "date", "meeting_date"))
        title = str(value(values, "title", "event_title", "meeting", "agenda_item") or "").strip()
        event_type = str(value(values, "type", "event_type", "meeting_type", "category") or "").strip()
        if event_date is None or not title or not event_type:
            continue
        prep = str(
            value(values, "prep_needed", "preparation", "prep", "notes_agenda", "notes", "agenda_notes")
            or "No preparation request was supplied."
        ).strip()
        related_bu = str(value(values, "related_bu", "business_unit") or "").strip()
        is_business_relevant, relevance_reason = _business_relevance(
            explicit_value=value(
                values,
                "business_relevant",
                "is_business",
                "relevance",
                "calendar_relevance",
                "visibility",
            ),
            title=title,
            event_type=event_type,
            related_bu=related_bu,
        )
        if not is_business_relevant:
            excluded_count += 1
            continue
        items.append({
            "event_id": f"calendar-{event_date.isoformat()}-{len(items) + 1}",
            "date": event_date.isoformat(),
            "day": event_date.strftime("%a %d %b"),
            "when": str(value(values, "start_time", "start", "time") or event_date.isoformat()),
            "ends_at": str(value(values, "end_time", "end") or "").strip() or None,
            "title": title,
            "type": event_type,
            "prep": prep,
            "attendees": str(value(values, "attendees", "participants") or "").strip() or None,
            "location": str(value(values, "location", "venue") or "").strip() or None,
            "related_bu": related_bu or None,
            "business_relevant": True,
            "relevance_reason": relevance_reason,
            "evidence_scope": "calendar_agenda_only" if restricted else "governed_calendar",
        })
    items.sort(key=lambda item: str(item.get("date") or ""))
    projection_day = workbook_anchor or date.today()
    projection_end = projection_day + timedelta(days=7)
    future_items = [item for item in items if str(item.get("date") or "") >= projection_day.isoformat()]
    projected_items = [
        item
        for item in future_items
        if str(item.get("date") or "") <= projection_end.isoformat()
    ]
    projection_policy = "next_7_days"
    return {
        "status": "ready" if items else "unavailable",
        "items": projected_items,
        "total_item_count": len(items),
        "excluded_non_business_count": excluded_count,
        "upcoming_item_count": len(projected_items),
        "future_item_count": len(future_items),
        "projection_as_of": projection_day.isoformat(),
        "projection_through": projection_end.isoformat(),
        "projection_policy": projection_policy,
        "projection_anchor_source": "workbook_today_anchor" if workbook_anchor else "system_date",
        "reason": (
            "No calendar item has been classified as business-relevant for the CEO projection."
            if not items and excluded_count
            else "Calendar workbook contains no complete Event_Date, Title and Type rows."
            if not items
            else "No business-relevant commitment falls within the next seven days."
            if not projected_items
            else None
        ),
        "source_file": relative_source,
        "sheet": sheet.title,
        "restricted": restricted,
        "evidence_scope": "calendar_agenda_only" if restricted else "governed_calendar",
    }


def _header_key(value: Any) -> str:
    return re.sub(r"[^a-z0-9]+", "_", str(value or "").strip().casefold()).strip("_")


def _date(value: Any) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    for pattern in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(str(value), pattern).date()
        except (TypeError, ValueError):
            continue
    return None
