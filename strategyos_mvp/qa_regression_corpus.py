from __future__ import annotations

import csv
import json
from pathlib import Path
from typing import Any, Callable, Iterable, Mapping


def load_ceo_questions(path: Path) -> list[dict[str, Any]]:
    """Load a delivered CEO-question corpus without assuming its business vocabulary."""
    suffix = path.suffix.casefold()
    if suffix == ".json":
        payload = json.loads(path.read_text(encoding="utf-8"))
        rows = payload if isinstance(payload, list) else payload.get("questions", [])
    elif suffix == ".jsonl":
        rows = [json.loads(line) for line in path.read_text(encoding="utf-8").splitlines() if line.strip()]
    elif suffix == ".csv":
        with path.open(encoding="utf-8-sig", newline="") as handle:
            rows = list(csv.DictReader(handle))
    elif suffix == ".xlsx":
        from openpyxl import load_workbook

        sheet = load_workbook(path, read_only=True, data_only=True).active
        values = list(sheet.iter_rows(values_only=True))
        headers = [str(value or "").strip() for value in values[0]] if values else []
        rows = [dict(zip(headers, row)) for row in values[1:]]
    else:
        raise ValueError(f"Unsupported CEO-question corpus format: {path.suffix}")

    normalized: list[dict[str, Any]] = []
    for index, row in enumerate(rows, start=1):
        item = dict(row) if isinstance(row, Mapping) else {"question": row}
        normalized_keys = {
            str(key or "").strip().casefold().replace(" ", "_"): value
            for key, value in item.items()
        }
        question = str(
            normalized_keys.get("question")
            or normalized_keys.get("prompt")
            or ""
        ).strip()
        if not question:
            continue
        normalized.append(
            {
                **item,
                "id": str(
                    normalized_keys.get("id")
                    or normalized_keys.get("#")
                    or f"question-{index}"
                ),
                "question": question,
            }
        )
    return normalized


def run_ceo_question_corpus(
    questions: Iterable[Mapping[str, Any]],
    answer: Callable[[str], Mapping[str, Any]],
) -> dict[str, Any]:
    """Run the delivered corpus and expose contract-level scoring for review."""
    rows: list[dict[str, Any]] = []
    valid_tiers = {"governed_fact", "derived_insight", "advisory"}
    for item in questions:
        payload = dict(answer(str(item["question"])))
        tier = str(payload.get("determinism_tier") or "")
        rows.append(
            {
                "id": str(item.get("id") or ""),
                "question": str(item["question"]),
                "answer": str(payload.get("answer") or ""),
                "matched": bool(payload.get("matched")),
                "determinism_tier": tier,
                "answered": bool(str(payload.get("answer") or "").strip()),
                "tier_visible": tier in valid_tiers,
                "derivability_violation": bool(payload.get("derivability_violation")),
                "missing_layer": payload.get("missing_layer"),
            }
        )
    return {
        "question_count": len(rows),
        "answered_count": sum(1 for row in rows if row["answered"]),
        "tiered_count": sum(1 for row in rows if row["tier_visible"]),
        "derivability_violation_count": sum(1 for row in rows if row["derivability_violation"]),
        "items": rows,
    }
