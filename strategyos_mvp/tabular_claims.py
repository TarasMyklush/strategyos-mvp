"""Explicit, versioned per-column/row semantics for mixed tabular sources.

The operator supplies the mapping. Neither filenames nor adjacent columns can
promote an ambiguous value to an actual. This module performs no I/O or actions.
"""
from datetime import date
from decimal import Decimal, InvalidOperation
from typing import Any

from pydantic import BaseModel, ConfigDict, Field, model_validator

from .source_claims import ClaimDraft, ClaimKind, explicit_claim_kind

INTERPRETER_VERSION = "2"


class ColumnClaimMapping(BaseModel):
    model_config = ConfigDict(extra="forbid")
    column: str = Field(min_length=1, max_length=160)
    metric_key: str = Field(min_length=1, max_length=240)
    claim_kind: ClaimKind | None = None
    kind_column: str | None = Field(default=None, min_length=1, max_length=160)
    unit: str = Field(min_length=1, max_length=80)
    scale: Decimal = Field(gt=0, allow_inf_nan=False)
    currency: str | None = Field(default=None, min_length=3, max_length=3)
    period_start: date | None = None
    period_end: date | None = None
    author_identity: str | None = Field(default=None, max_length=240)
    author_column: str | None = Field(default=None, max_length=160)

    @model_validator(mode="after")
    def explicit_kind_source(self):
        if (self.claim_kind is None) == (self.kind_column is None):
            raise ValueError("Choose exactly one explicit claim kind or per-row kind column.")
        if self.period_start and self.period_end and self.period_end < self.period_start:
            raise ValueError("Column period end precedes its start.")
        return self


class TableClaimMapping(BaseModel):
    model_config = ConfigDict(extra="forbid")
    mapping_key: str = Field(pattern=r"^[A-Za-z0-9][A-Za-z0-9_.:-]{0,159}$")
    mapping_version: str = Field(min_length=1, max_length=80)
    rationale: str = Field(min_length=1, max_length=2000)
    sheet: str = Field(min_length=1, max_length=160)
    subject_type: str = Field(min_length=1, max_length=80)
    subject_key_column: str = Field(min_length=1, max_length=160)
    business_unit_column: str | None = Field(default=None, max_length=160)
    period_start_column: str | None = Field(default=None, max_length=160)
    period_end_column: str | None = Field(default=None, max_length=160)
    scenario_column: str | None = Field(default=None, max_length=160)
    columns: list[ColumnClaimMapping] = Field(min_length=1, max_length=50)

    @model_validator(mode="after")
    def unique_columns(self):
        if len({item.column for item in self.columns}) != len(self.columns):
            raise ValueError("A source column can have only one interpretation per mapping.")
        return self


def read_workbook_rows(content: bytes, mapping: TableClaimMapping) -> list[dict[str, Any]]:
    from pathlib import Path
    from tempfile import NamedTemporaryFile
    from openpyxl import load_workbook
    from .review_files import _validate_ooxml
    with NamedTemporaryFile(suffix=".xlsx") as uploaded:
        uploaded.write(content)
        uploaded.flush()
        _validate_ooxml(Path(uploaded.name), ".xlsx")
        book = load_workbook(uploaded.name, data_only=True, read_only=True, keep_links=False)
        try:
            if mapping.sheet not in book.sheetnames:
                raise ValueError("The mapped worksheet is not present.")
            sheet = book[mapping.sheet]
            limit = 500 // len(mapping.columns)
            if sheet.max_row is not None and sheet.max_row > limit + 1:
                raise ValueError("The worksheet exceeds the 500-cell mapping limit; use bounded source batches.")
            values = list(sheet.iter_rows(max_row=limit + 2, max_col=200, values_only=True))
            header = [str(value).strip() if value is not None else "" for value in values[0]]
            named = [name for name in header if name]
            if len(set(named)) != len(named):
                raise ValueError("Duplicate worksheet headers require explicit source correction.")
            required = {mapping.subject_key_column, *(column.column for column in mapping.columns)}
            required.update(key for key in (mapping.business_unit_column, mapping.period_start_column,
                mapping.period_end_column, mapping.scenario_column) if key)
            required.update(key for column in mapping.columns for key in (column.kind_column, column.author_column) if key)
            if required.difference(header):
                raise ValueError("The mapping refers to missing worksheet columns.")
            if any(value is not None for value in values[-1]) and len(values) > limit + 1:
                raise ValueError("The worksheet exceeds the 500-cell mapping limit; use bounded source batches.")
            rows = []
            body = values[1:limit + 1]
            while body and not any(value is not None for value in body[-1]):
                body.pop()
            for values_row in body:
                rows.append({key: (value.date().isoformat() if hasattr(value, "date") else value)
                             for key, value in zip(header, values_row) if key})
            return rows
        finally:
            book.close()


def map_table(rows: list[dict[str, Any]], mapping: TableClaimMapping, *,
              tenant_id: str, source_key: str, occurrence_key: str, recorded_by: str) -> dict[str, Any]:
    if len(rows) * len(mapping.columns) > 500:
        raise ValueError("A mapping batch is limited to 500 source cells.")
    drafts, issues = [], []
    missing = 0
    for index, row in enumerate(rows, start=2):
        subject_value = row.get(mapping.subject_key_column)
        subject = str(subject_value).strip() if subject_value is not None else ""
        for column in mapping.columns:
            locator = f"{mapping.sheet}!row {index}; column {column.column}"
            raw = row.get(column.column)
            reasons = []
            if raw is None or raw == "":
                missing += 1
                issues.append({"locator": locator, "reason": "value_missing", "disposition": "no_claim"})
                continue
            if not subject:
                issues.append({"locator": locator, "reason": "subject_unresolved", "disposition": "no_claim"})
                continue
            kind = column.claim_kind if column.claim_kind is not None else explicit_claim_kind(row.get(column.kind_column))
            if kind == ClaimKind.UNKNOWN:
                reasons.append("claim_kind_ambiguous")
            try:
                if isinstance(raw, bool):
                    raise ValueError("Boolean is not a financial number")
                value = Decimal(str(raw))
                if not value.is_finite():
                    raise ValueError("Nonfinite value")
            except (InvalidOperation, ValueError):
                value = None
                reasons.append("numeric_value_invalid")
            def period(fixed: date | None, key: str | None) -> date | None:
                if fixed is not None:
                    return fixed
                supplied = row.get(key) if key else None
                if not supplied:
                    return None
                return date.fromisoformat(str(supplied))
            try:
                start = period(column.period_start, mapping.period_start_column)
                end = period(column.period_end, mapping.period_end_column)
                if start is None or end is None or end < start:
                    raise ValueError("Period unresolved")
            except ValueError:
                start = end = None
                reasons.append("period_unresolved")
            author = column.author_identity or (str(row.get(column.author_column) or "").strip() if column.author_column else None)
            if kind == ClaimKind.FORECAST and not author:
                reasons.append("forecast_author_missing")
            if reasons:
                kind = ClaimKind.UNKNOWN
                issues.extend({"locator": locator, "reason": reason, "disposition": "quarantined"} for reason in reasons)
            business_unit = str(row.get(mapping.business_unit_column) or "").strip() or None
            if mapping.subject_type in {"business_unit", "bu"} and not mapping.business_unit_column:
                business_unit = subject
            drafts.append(ClaimDraft(tenant_id=tenant_id,
                assertion_namespace=f"table:{source_key}:{mapping.mapping_key}",
                subject_type=mapping.subject_type, subject_key=subject,
                metric_key=column.metric_key, claim_kind=kind, production_method="extracted",
                value_numeric=value, value_text=str(raw) if value is None else None,
                unit=column.unit, scale=column.scale, currency=column.currency,
                business_unit=business_unit,
                period_start=start, period_end=end, author_identity=author,
                scenario_key=str(row.get(mapping.scenario_column) or "").strip() or None,
                dimensions={"source_column": column.column, "source_sheet": mapping.sheet},
                source_occurrence_keys=(occurrence_key,),
                metadata={"mapping_key": mapping.mapping_key, "mapping_version": mapping.mapping_version,
                    "mapping_engine_version": INTERPRETER_VERSION,
                    "mapping_rationale": mapping.rationale, "recorded_by": recorded_by,
                    "source_locator": locator, "quarantine_reasons": reasons}))
    return {"drafts": drafts, "issues": issues, "mapping_engine_version": INTERPRETER_VERSION,
            "source_cell_count": len(rows) * len(mapping.columns),
            "claim_count": len(drafts), "missing_count": missing,
            "quarantined_count": sum(draft.claim_kind == ClaimKind.UNKNOWN for draft in drafts),
            "unmapped_count": len(rows) * len(mapping.columns) - len(drafts)}
