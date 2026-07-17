from __future__ import annotations

import re
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string, get_column_letter

from .source_governance import CONTROL_PLANE, CURRENT_EVIDENCE, EVALUATOR_ONLY


_FORMULA_ERROR_RE = re.compile(r"#(?:REF!|DIV/0!|VALUE!|NAME\?|N/A|NUM!|NULL!)", re.IGNORECASE)
_SIMPLE_SUM_RE = re.compile(
    r"^=SUM\(\$?([A-Z]{1,3})\$?(\d+):\$?([A-Z]{1,3})\$?(\d+)\)$",
    re.IGNORECASE,
)
_CELL_REF_RE = re.compile(r"(?<![A-Z0-9_])\$?([A-Z]{1,3})\$?(\d+)", re.IGNORECASE)
_TOTAL_DUE_RE = re.compile(
    r"(?:TOTAL\s+(?:TTC\s*/\s*)?TOTAL\s+DUE|TOTAL\s+DUE|AMOUNT\s+DUE)"
    r".{0,100}?([A-Z]{3})\s*([\d,]+(?:\.\d{1,2})?)",
    re.IGNORECASE | re.DOTALL,
)
_VENDOR_STOP_WORDS = {
    "company",
    "distribution",
    "international",
    "limited",
    "pharma",
    "pharmaceutical",
    "pharmaceuticals",
    "services",
    "supplier",
    "trading",
}


def _issue(
    *,
    code: str,
    severity: str,
    source_file: str,
    detail: str,
    sheet: str | None = None,
    cell: str | None = None,
    formula: str | None = None,
    metrics: dict[str, Any] | None = None,
) -> dict[str, Any]:
    payload: dict[str, Any] = {
        "code": code,
        "severity": severity,
        "source_file": source_file,
        "detail": detail,
    }
    if sheet:
        payload["sheet"] = sheet
    if cell:
        payload["cell"] = cell
    if formula:
        payload["formula"] = formula
    if metrics:
        payload["metrics"] = metrics
    return payload


def _relative_formula_shape(formula: str, formula_row: int) -> str:
    return _CELL_REF_RE.sub(
        lambda match: f"{match.group(1).upper()}[{int(match.group(2)) - formula_row:+d}]",
        formula.upper().replace("$", ""),
    )


def inspect_workbook_formula_quality(path: Path, *, source_file: str | None = None) -> list[dict[str, Any]]:
    """Detect broken and locally inconsistent formulas without changing a workbook."""

    source_name = source_file or path.name
    try:
        workbook = load_workbook(path, data_only=False, read_only=False)
    except Exception as exc:
        return [
            _issue(
                code="workbook_unreadable",
                severity="high",
                source_file=source_name,
                detail=f"Workbook formula inspection failed: {exc}",
            )
        ]

    issues: list[dict[str, Any]] = []
    seen: set[tuple[str, str, str]] = set()
    for sheet in workbook.worksheets:
        formulas_by_column: dict[int, dict[int, str]] = {}
        for row in sheet.iter_rows():
            for cell in row:
                formula = cell.value
                if not isinstance(formula, str) or not formula.startswith("="):
                    continue
                formulas_by_column.setdefault(cell.column, {})[cell.row] = formula
                if _FORMULA_ERROR_RE.search(formula):
                    key = ("formula_error_token", sheet.title, cell.coordinate)
                    if key not in seen:
                        seen.add(key)
                        issues.append(
                            _issue(
                                code="formula_error_token",
                                severity="high",
                                source_file=source_name,
                                sheet=sheet.title,
                                cell=cell.coordinate,
                                formula=formula,
                                detail="Formula contains an Excel error token and cannot calculate reliably.",
                            )
                        )

                match = _SIMPLE_SUM_RE.match(formula.replace(" ", ""))
                if match and int(match.group(2)) == cell.row and int(match.group(4)) == cell.row:
                    end_column = column_index_from_string(match.group(3))
                    omitted_column = cell.column - 1
                    omitted_cell = sheet.cell(row=cell.row, column=omitted_column)
                    if end_column == omitted_column - 1 and omitted_cell.value not in {None, ""}:
                        key = ("sum_omits_adjacent_cell", sheet.title, cell.coordinate)
                        if key not in seen:
                            seen.add(key)
                            issues.append(
                                _issue(
                                    code="sum_omits_adjacent_cell",
                                    severity="high",
                                    source_file=source_name,
                                    sheet=sheet.title,
                                    cell=cell.coordinate,
                                    formula=formula,
                                    detail=(
                                        "SUM range stops before the immediately adjacent populated cell "
                                        f"{get_column_letter(omitted_column)}{cell.row}."
                                    ),
                                )
                            )

                referenced_rows = [int(match.group(2)) for match in _CELL_REF_RE.finditer(formula)]
                if (
                    len(referenced_rows) >= 2
                    and cell.row in referenced_rows
                    and any(row_no != cell.row for row_no in referenced_rows)
                    and ":" not in formula
                ):
                    key = ("inconsistent_relative_formula", sheet.title, cell.coordinate)
                    if key not in seen:
                        seen.add(key)
                        issues.append(
                            _issue(
                                code="inconsistent_relative_formula",
                                severity="high",
                                source_file=source_name,
                                sheet=sheet.title,
                                cell=cell.coordinate,
                                formula=formula,
                                detail=(
                                    "Formula mixes a current-row reference with a different-row reference; "
                                    "confirm that the cross-row dependency is intentional."
                                ),
                            )
                        )

        for column, formulas in formulas_by_column.items():
            for row_no, formula in formulas.items():
                previous = formulas.get(row_no - 1)
                following = formulas.get(row_no + 1)
                if previous is None or following is None:
                    continue
                previous_shape = _relative_formula_shape(previous, row_no - 1)
                following_shape = _relative_formula_shape(following, row_no + 1)
                current_shape = _relative_formula_shape(formula, row_no)
                if previous_shape != following_shape or current_shape == previous_shape:
                    continue
                coordinate = f"{get_column_letter(column)}{row_no}"
                key = ("inconsistent_relative_formula", sheet.title, coordinate)
                if (
                    ("sum_omits_adjacent_cell", sheet.title, coordinate) in seen
                    or ("formula_error_token", sheet.title, coordinate) in seen
                ):
                    continue
                if key in seen:
                    continue
                seen.add(key)
                issues.append(
                    _issue(
                        code="inconsistent_relative_formula",
                        severity="high",
                        source_file=source_name,
                        sheet=sheet.title,
                        cell=coordinate,
                        formula=formula,
                        detail="Formula reference pattern differs from the matching formulas immediately above and below.",
                    )
                )
    workbook.close()
    return issues


def _current_role_path(manifest: list[dict[str, Any]], role: str) -> tuple[str, Path] | None:
    matches: list[tuple[str, Path]] = []
    for item in manifest:
        classification = item.get("classification") or {}
        if str(item.get("source_disposition") or CURRENT_EVIDENCE) != CURRENT_EVIDENCE:
            continue
        if classification.get("status") != "classified" or classification.get("role") != role:
            continue
        normalized_path = item.get("normalized_path")
        if normalized_path:
            matches.append((str(item.get("relative_path") or ""), Path(str(normalized_path))))
    return matches[0] if len(matches) == 1 else None


def _read_frame(path: Path) -> pd.DataFrame:
    if path.suffix.lower() == ".csv":
        return pd.read_csv(path)
    return pd.read_excel(path)


def _balance_issue(
    manifest: list[dict[str, Any]],
    *,
    role: str,
    debit_column: str,
    credit_column: str,
) -> dict[str, Any] | None:
    resolved = _current_role_path(manifest, role)
    if resolved is None:
        return None
    source_file, path = resolved
    try:
        frame = _read_frame(path)
        debit = float(pd.to_numeric(frame[debit_column], errors="coerce").fillna(0).sum())
        credit = float(pd.to_numeric(frame[credit_column], errors="coerce").fillna(0).sum())
    except Exception as exc:
        return _issue(
            code=f"{role}_reconciliation_unreadable",
            severity="high",
            source_file=source_file,
            detail=f"Could not perform debit/credit reconciliation: {exc}",
        )
    difference = debit - credit
    if abs(difference) <= 0.01:
        return None
    label = "General ledger" if role == "gl_extract" else "Trial balance"
    return _issue(
        code=f"{role}_debit_credit_imbalance",
        severity="medium",
        source_file=source_file,
        detail=f"{label} debit and credit totals do not reconcile within SAR 0.01.",
        metrics={
            "debit_total_sar": round(debit, 2),
            "credit_total_sar": round(credit, 2),
            "difference_sar": round(difference, 2),
        },
    )


def _document_text(item: dict[str, Any], raw_root: Path) -> str:
    extraction = item.get("text_extraction") or {}
    raw_text = str(extraction.get("raw_text") or "")
    if raw_text:
        return raw_text
    path = raw_root / str(item.get("relative_path") or "")
    if path.suffix.lower() in {".txt", ".md"}:
        try:
            return path.read_text(encoding="utf-8", errors="ignore")
        except OSError:
            return ""
    return ""


def _vendor_anchor(vendor_name: str) -> str | None:
    candidates = [
        token.lower()
        for token in re.findall(r"[A-Za-z]{4,}", vendor_name)
        if token.lower() not in _VENDOR_STOP_WORDS
    ]
    return candidates[0] if candidates else None


def _reference_consistency_issues(
    manifest: list[dict[str, Any]],
    raw_root: Path,
) -> list[dict[str, Any]]:
    """Cross-check non-SAR AP references against invoices and correspondence."""

    resolved = _current_role_path(manifest, "ap_ledger")
    if resolved is None:
        return []
    source_file, ap_path = resolved
    try:
        ap = _read_frame(ap_path)
    except Exception:
        return []
    required = {"Invoice_ID", "Vendor_Name", "Currency", "Amount_Original_Currency"}
    if not required.issubset(ap.columns):
        return []

    documents: list[tuple[dict[str, Any], str]] = []
    for item in manifest:
        role = str((item.get("classification") or {}).get("role") or "")
        if role not in {"invoice_document", "email_correspondence"}:
            continue
        text = _document_text(item, raw_root)
        if text:
            documents.append((item, text))

    issues: list[dict[str, Any]] = []
    non_sar = ap.loc[ap["Currency"].astype(str).str.upper().ne("SAR")]
    for _, row in non_sar.iterrows():
        invoice_id = str(row.get("Invoice_ID") or "").strip()
        vendor_name = str(row.get("Vendor_Name") or "").strip()
        currency = str(row.get("Currency") or "").strip().upper()
        try:
            ledger_amount = float(row.get("Amount_Original_Currency"))
        except (TypeError, ValueError):
            continue
        if not invoice_id:
            continue
        vendor_anchor = _vendor_anchor(vendor_name)
        for item, text in documents:
            reference_index = text.lower().find(invoice_id.lower())
            if reference_index < 0:
                continue
            role = str((item.get("classification") or {}).get("role") or "")
            evidence_file = str(item.get("relative_path") or "")
            if role == "email_correspondence" and vendor_anchor:
                window = text[max(0, reference_index - 180) : reference_index + 260].lower()
                if vendor_anchor not in window:
                    issues.append(
                        _issue(
                            code="reference_vendor_identity_conflict",
                            severity="high",
                            source_file=evidence_file,
                            detail=(
                                f"Reference {invoice_id} is linked to ledger vendor '{vendor_name}', "
                                "but the nearby correspondence identity does not name that vendor."
                            ),
                            metrics={
                                "invoice_id": invoice_id,
                                "ledger_vendor": vendor_name,
                                "ledger_source_file": source_file,
                            },
                        )
                    )
            if role != "invoice_document":
                continue
            total_match = _TOTAL_DUE_RE.search(text)
            if not total_match or total_match.group(1).upper() != currency:
                continue
            document_amount = float(total_match.group(2).replace(",", ""))
            if abs(document_amount - ledger_amount) <= 0.01:
                continue
            issues.append(
                _issue(
                    code="invoice_ledger_amount_conflict",
                    severity="high",
                    source_file=evidence_file,
                    detail=(
                        f"Invoice {invoice_id} total does not match the AP ledger's original-currency amount."
                    ),
                    metrics={
                        "invoice_id": invoice_id,
                        "currency": currency,
                        "ledger_amount": round(ledger_amount, 2),
                        "document_amount": round(document_amount, 2),
                        "difference": round(document_amount - ledger_amount, 2),
                        "ledger_source_file": source_file,
                    },
                )
            )
    return issues


def build_source_quality_report(manifest: list[dict[str, Any]], raw_root: Path) -> dict[str, Any]:
    issues: list[dict[str, Any]] = []
    inspected_workbooks = 0
    for item in manifest:
        if not item.get("supported") or str(item.get("source_disposition")) in {
            CONTROL_PLANE,
            EVALUATOR_ONLY,
        }:
            continue
        rel = str(item.get("relative_path") or "")
        if Path(rel).suffix.lower() != ".xlsx":
            continue
        inspected_workbooks += 1
        issues.extend(inspect_workbook_formula_quality(raw_root / rel, source_file=rel))

    for reconciliation in (
        _balance_issue(
            manifest,
            role="gl_extract",
            debit_column="Debit",
            credit_column="Credit",
        ),
        _balance_issue(
            manifest,
            role="trial_balance",
            debit_column="Debit_Total",
            credit_column="Credit_Total",
        ),
    ):
        if reconciliation:
            issues.append(reconciliation)
    issues.extend(_reference_consistency_issues(manifest, raw_root))

    counts: dict[str, int] = {}
    for item in issues:
        severity = str(item.get("severity") or "unknown")
        counts[severity] = counts.get(severity, 0) + 1
    status = "review_required" if counts.get("high") else "warning" if issues else "clean"
    return {
        "status": status,
        "inspected_workbook_count": inspected_workbooks,
        "issue_count": len(issues),
        "counts_by_severity": counts,
        "issues": issues,
    }


def build_acceptance_readiness(
    manifest: list[dict[str, Any]],
    *,
    file_accounting: dict[str, Any],
    source_quality: dict[str, Any],
) -> dict[str, Any]:
    failed_pdfs = [
        str(item.get("relative_path") or "")
        for item in manifest
        if item.get("file_type_hint") == "pdf" and item.get("extraction_status") != "ok"
    ]
    control_plane_leaks = [
        str(item.get("relative_path") or "")
        for item in manifest
        if item.get("source_disposition") in {CONTROL_PLANE, EVALUATOR_ONLY}
        and item.get("normalized_path")
    ]
    blocking_reasons: list[str] = []
    review_reasons: list[str] = []
    if file_accounting.get("silent_omission_count"):
        blocking_reasons.append("One or more source files have no governed processing disposition.")
    if control_plane_leaks:
        blocking_reasons.append("Control-plane or evaluator-only files entered normalized run evidence.")
    if failed_pdfs:
        review_reasons.append(f"{len(failed_pdfs)} PDF files did not complete text extraction.")
    high_quality_issues = int((source_quality.get("counts_by_severity") or {}).get("high") or 0)
    medium_quality_issues = int((source_quality.get("counts_by_severity") or {}).get("medium") or 0)
    if high_quality_issues:
        review_reasons.append(f"{high_quality_issues} high-severity source quality issues require review.")
    if medium_quality_issues:
        review_reasons.append(f"{medium_quality_issues} reconciliation warnings require a declared reporting basis.")
    status = "blocked" if blocking_reasons else "review_required" if review_reasons else "ready"
    return {
        "status": status,
        "ready_for_acceptance": status == "ready",
        "blocking_reasons": blocking_reasons,
        "review_reasons": review_reasons,
        "failed_pdf_extractions": failed_pdfs,
        "control_plane_evidence_leaks": control_plane_leaks,
    }
