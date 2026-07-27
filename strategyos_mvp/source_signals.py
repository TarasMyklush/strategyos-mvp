from __future__ import annotations

import re
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


def _normal(value: Any) -> str:
    return "".join(character for character in str(value or "").lower() if character.isalnum())


def _tone(signal_type: str) -> str:
    kinds = {
        part
        for part in re.split(r"[^a-z]+", str(signal_type or "").lower())
        if part
    }
    if "threat" in kinds and "opportunity" in kinds:
        return "watch"
    if "threat" in kinds:
        return "critical"
    if "opportunity" in kinds:
        return "positive"
    return "watch"


def derive_governed_signals(dataset_root: Path, *, cap: int = 12) -> dict[str, Any]:
    """Discover a governed signal register by schema and expose bounded rows."""
    root = Path(dataset_root)
    required = {
        "signalid",
        "detected",
        "type",
        "affectedbus",
        "signal",
        "potentialimpact",
        "probability",
        "horizon",
        "recommendedaction",
    }
    for path in sorted(root.rglob("*.xlsx")):
        try:
            book = load_workbook(path, read_only=True, data_only=True)
        except Exception:
            continue
        for sheet in book.worksheets:
            rows = sheet.iter_rows(values_only=True)
            header_row = next(rows, None)
            headers = {
                _normal(value): index
                for index, value in enumerate(header_row or ())
                if value is not None
            }
            if not required.issubset(headers):
                continue

            def cell(values: tuple[Any, ...], name: str) -> Any:
                index = headers.get(name)
                return values[index] if index is not None and index < len(values) else None

            items: list[dict[str, Any]] = []
            for values in rows:
                values = tuple(values)
                signal_id = str(cell(values, "signalid") or "").strip()
                statement = str(cell(values, "signal") or "").strip()
                if not signal_id or not statement:
                    continue
                signal_type = str(cell(values, "type") or "").strip()
                source = str(cell(values, "source") or "").strip()
                source_type = str(cell(values, "sourcetype") or "").strip()
                items.append(
                    {
                        "key": signal_id,
                        "title": statement,
                        "summary": str(cell(values, "potentialimpact") or "").strip(),
                        "tone": _tone(signal_type),
                        "classification": " · ".join(
                            part
                            for part in (
                                signal_type,
                                str(cell(values, "affectedbus") or "").strip(),
                            )
                            if part
                        ),
                        "action_required": False,
                        "context": {
                            "what": statement,
                            "why_attached": str(cell(values, "potentialimpact") or "").strip(),
                            "sources": [
                                value
                                for value in (source_type, source)
                                if value
                            ],
                            "detected": str(cell(values, "detected") or "").strip(),
                            "probability": str(cell(values, "probability") or "").strip(),
                            "horizon": str(cell(values, "horizon") or "").strip(),
                            "leading_indicator": str(
                                cell(values, "leadingindicatortowatch") or ""
                            ).strip(),
                            "recommended_action": str(
                                cell(values, "recommendedaction") or ""
                            ).strip(),
                        },
                    }
                )
            if items:
                priority = {"critical": 0, "watch": 1, "positive": 2}
                items.sort(
                    key=lambda item: (
                        priority.get(str(item.get("tone")), 3),
                        str(item.get("key")),
                    )
                )
                return {
                    "status": "ready",
                    "cap": cap,
                    "total_item_count": len(items),
                    "items": items[:cap],
                    "source_file": str(path.relative_to(root)),
                    "sheet": sheet.title,
                }
    return {
        "status": "unavailable",
        "cap": cap,
        "total_item_count": 0,
        "items": [],
        "reason": "No governed signal-register schema is present in the run model.",
    }
