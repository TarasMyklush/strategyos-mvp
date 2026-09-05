"""Index approved source rows/pages with exact citations into the scoped collection."""
from pathlib import Path
from itertools import islice
import logging
logger = logging.getLogger(__name__)
from . import semantic_embeddings, vector_store
from .evidence import sha256_file

MAX_FILE_ROWS = 10000
MAX_CHUNKS = 50000


def source_records(evidence):
    root = evidence.dataset_root.resolve()
    for relative, entry in sorted(evidence.manifest.items()):
        from .source_governance import initial_source_disposition, CONTROL_PLANE, EVALUATOR_ONLY
        if initial_source_disposition(relative) in {CONTROL_PLANE, EVALUATOR_ONLY}:
            continue
        path = (root / relative).resolve()
        if not path.is_relative_to(root) or not path.is_file() or sha256_file(path) != entry['sha256']:
            raise ValueError('Source changed after ingestion; semantic indexing was stopped.')
        if relative in evidence.pdf_text:
            for page, text in enumerate(evidence.pdf_text[relative], start=1):
                for chunk in vector_store._chunk_text(text):
                    yield relative, entry['sha256'], f'PDF page {page}', chunk
        elif path.suffix.lower() in {'.txt', '.md'}:
            text = path.read_text(encoding='utf-8-sig')
            if len(text) > 2_000_000:
                raise ValueError(f'{relative}: text exceeds reviewed indexing capacity.')
            for number, chunk in enumerate(vector_store._chunk_text(text), start=1):
                yield relative, entry['sha256'], f'Text chunk {number}', chunk
        elif path.suffix.lower() == '.docx':
            from zipfile import ZipFile
            from xml.etree import ElementTree
            with ZipFile(path) as archive:
                info = archive.getinfo('word/document.xml')
                if info.file_size > 2_000_000:
                    raise ValueError(f'{relative}: document exceeds reviewed indexing capacity.')
                document = ElementTree.fromstring(archive.read(info))
            namespace = '{http://schemas.openxmlformats.org/wordprocessingml/2006/main}'
            for number, paragraph in enumerate(document.iter(namespace + 'p'), start=1):
                text = ''.join(node.text or '' for node in paragraph.iter(namespace + 't'))
                for chunk in vector_store._chunk_text(text):
                    yield relative, entry['sha256'], f'Document paragraph {number}', chunk
        elif path.suffix.lower() == '.xlsx':
            from openpyxl import load_workbook
            book = load_workbook(path, read_only=True, data_only=True)
            try:
                for sheet in book:
                    if sheet.max_row is not None and sheet.max_row > MAX_FILE_ROWS:
                        raise ValueError(f'{relative}: worksheet exceeds the reviewed indexing capacity.')
                    rows = sheet.iter_rows(values_only=True)
                    headers = [str(x or '') for x in next(rows, ())]
                    header_keys = {header.strip().casefold() for header in headers}
                    if 'question' in header_keys and {'answer type', 'where strategyos finds the answer'} & header_keys:
                        continue
                    for number, values in enumerate(rows, start=2):
                        if number > MAX_FILE_ROWS:
                            raise ValueError(f"{relative}: worksheet exceeds the reviewed indexing capacity.")
                        text = '; '.join(f'{headers[i] if i < len(headers) else i}: {value}' for i, value in enumerate(values) if value is not None)
                        if text:
                            yield relative, entry['sha256'], f'{sheet.title}!Excel row {number}', text[:vector_store.MAX_INDEX_TEXT]
            finally:
                book.close()


def sync_sources(*, run_id, tenant_slug, evidence):
    if not semantic_embeddings.configured():
        return {'status': 'disabled', 'reason': 'No pinned local embedding model configured.'}
    if evidence is None:
        raise ValueError("Source evidence is unavailable; indexing cannot continue.")
    vector_store._run_filter(run_id)  # Reject unknown or inaccessible runs before reading source files.
    vector_store._ensure_collection()
    # Preflight the entire bounded manifest before writing any index records.
    records = list(islice(source_records(evidence), MAX_CHUNKS + 1))
    if len(records) > MAX_CHUNKS:
        raise ValueError('Source pack exceeds reviewed semantic indexing capacity.')
    allowed_paths = sorted({record[0] for record in records})
    obsolete_filter = {'must': [{'key':key, 'match':{'value':value}} for key,value in
                       (('run_id',run_id),('tenant_slug',tenant_slug),('point_type','source_chunk'))]}
    if allowed_paths:
        obsolete_filter['must_not'] = [{'key':'source_path','match':{'any':allowed_paths}}]
    vector_store._qdrant_request('POST', f'/collections/{vector_store.COLLECTION_NAME}/points/delete?wait=true', {'filter':obsolete_filter})
    reused = 0
    for offset in range(0, len(records), 64):
        batch = []
        for relative, digest, locator, text in records[offset:offset + 64]:
            batch.append({'id': vector_store._point_id(run_id, 'source_chunk', relative, locator, text),
                          'payload': {'run_id': run_id, 'tenant_slug': tenant_slug, 'point_type': 'source_chunk',
                                      'source_path': relative, 'source_hash': digest, 'locator': locator,
                                      'title': Path(relative).stem, 'text': text, 'excerpt': text[:700]}})
        stored = vector_store._qdrant_request('POST', f'/collections/{vector_store.COLLECTION_NAME}/points',
                    {'ids': [point['id'] for point in batch], 'with_payload': ['run_id', 'tenant_slug', 'source_hash'], 'with_vector': False})
        existing = {str(point['id']): point.get('payload', {}) for point in stored.get('result', [])}
        missing = [point for point in batch if not all(existing.get(point['id'], {}).get(key) == point['payload'][key]
                   for key in ('run_id', 'tenant_slug', 'source_hash'))]
        reused += len(batch) - len(missing)
        vectors = semantic_embeddings.embed_many([point['payload']['title'] + ': ' + point['payload']['text'] for point in missing])
        for point, vector in zip(missing, vectors):
            point['vector'] = vector
        if missing:
            vector_store._qdrant_request('PUT', f'/collections/{vector_store.COLLECTION_NAME}/points?wait=true', {'points': missing})
        if offset % 1024 == 0:
            logger.info('Semantic source index: %s/%s records, %s reused', min(offset + 64, len(records)), len(records), reused)
    return {'status': 'ready', 'point_count': len(records), 'reused_points': reused,
            'collection': vector_store.COLLECTION_NAME, 'model_revision': semantic_embeddings.MODEL_REVISION}


def retrieve(run_id, question):
    if not semantic_embeddings.configured():
        return None
    result = vector_store.search_run_vectors(run_id, question, limit=12, point_type='source_chunk')
    if result.get('status') != 'ready':
        return {'status': 'unavailable', 'reason': result.get('reason')}
    return {'status': 'ready', 'records': [{key: row.get(key) for key in
            ('source_path', 'source_hash', 'locator', 'text', 'score')} for row in result.get('results', [])]}


def targeted_financial_records(evidence, question):
    """Complement semantic hits with bounded, explicitly named financial tables.

    A narrow topic contract prevents a near-neighbour search miss from being
    presented as absence of a connected payroll or peer-comparison workbook.
    """
    import re
    from types import SimpleNamespace
    text = str(question).casefold()
    patterns = []
    if re.search(r'\b(?:employee|employees|headcount|payroll|wage|wages)\b', text):
        patterns += ['headcount_payroll']
    if re.search(r'\bper employee\b', text):
        patterns += ['group_bu_pnl']
    if re.search(r'\b(?:peer|peers|competitor|competitors)\b', text):
        patterns += ['competitor_financials', 'group_bu_pnl']
    if re.search(r'\b(?:synergy|synergies|insourcing)\b', text):
        patterns += ['synergy_programme_charter', 'synergy_program_charter']
    customer_question = bool(re.search(r'\b(?:customer|customers|profitability|pricing|prices|terms)\b', text))
    revenue_question = bool(re.search(r'\b(?:revenue|growth|segment|segments)\b', text))
    if customer_question or revenue_question:
        patterns += ['revenue_analytics']
    manifest = {path: entry for path, entry in evidence.manifest.items()
                if path.lower().endswith(('.xlsx', '.docx')) and any(token in Path(path).stem.casefold() for token in patterns)}
    if not manifest:
        return {'status': 'not_applicable', 'records': []}
    scoped = SimpleNamespace(dataset_root=evidence.dataset_root, manifest=manifest, pdf_text={})
    records = []; characters = 0
    try:
        for path, digest, locator, content in source_records(scoped):
            if 'revenue_analytics' in Path(path).stem.casefold():
                customer_sheet = locator.startswith(('Customer_Profitability', 'Top_Customers'))
                if customer_sheet != customer_question:
                    continue
            if len(records) >= 96 or characters + len(content) > 30000:
                return {'status': 'bounded', 'coverage_complete': False, 'records': records}
            records.append({'source_path': path, 'source_hash': digest, 'locator': locator, 'text': content})
            characters += len(content)
    except (ValueError, OSError):
        return {'status': 'unavailable', 'reason': 'Required source bytes failed verification.', 'records': []}
    return {'status': 'ready', 'coverage_complete': True, 'records': records}
