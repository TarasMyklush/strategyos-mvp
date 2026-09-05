from __future__ import annotations

import json
import hashlib
import re
from pathlib import Path
from typing import Any

import pandas as pd

from .ingestion import DataBundle
from .models import Citation, Finding
from .prompt_injection import guard_untrusted_document_text
from .sensitive_ids import matches_sensitive_identifier_token


ROW_RE = re.compile(r"(?:Excel|CSV) row\s+(\d+)", re.I)
PAGE_RE = re.compile(r"page\s+(\d+)", re.I)


def canonical_workbook_locator(bundle: DataBundle, rel_path: str, locator: str) -> str | None:
    """Resolve an explicit record identifier to a unique hash-checked row."""
    if re.fullmatch(r'.+!Excel rows? \d+(?:-\d+)?', locator):
        return locator
    identifier = re.fullmatch(r'(?:(.+?)(?:!| record | ))?((?:SIG|KPI|EV|INIT|RD|CT)-[A-Za-z0-9-]+)', locator)
    business_unit = locator.removeprefix('Business unit / ') if locator.startswith('Business unit / ') else None
    if not identifier and not business_unit:
        return None
    entry = bundle.evidence.manifest.get(rel_path)
    root = bundle.evidence.dataset_root.resolve(); path = (root / rel_path).resolve()
    if not entry or not path.is_relative_to(root) or not path.is_file():
        return None
    if hashlib.sha256(path.read_bytes()).hexdigest() != entry.get('sha256'):
        return None
    from openpyxl import load_workbook
    book = load_workbook(path, read_only=True, data_only=True)
    matches = []
    try:
        for sheet in book:
            if identifier and identifier[1] and identifier[1] != sheet.title:
                continue
            rows = sheet.iter_rows(values_only=True)
            headers = next(rows, ())
            columns = [i for i, value in enumerate(headers)
                       if (re.sub(r'[^a-z0-9]', '', str(value).lower()) in {'businessunit', 'bu'} if business_unit
                           else re.sub(r'[^a-z0-9]', '', str(value).lower()).endswith('id'))]
            if not columns:
                continue
            for number, values in enumerate(rows, 2):
                if number > 10000:
                    return None
                if any(i < len(values) and str(values[i]).strip() == (business_unit or identifier[2]) for i in columns):
                    matches.append(f'{sheet.title}!Excel row {number}')
        return matches[0] if len(matches) == 1 else None
    finally:
        book.close()


def citation_location_hints(bundle: DataBundle, citations: list[dict[str, Any]]) -> list[dict[str, Any]]:
    """Supply candidate source rows for repair without silently choosing one."""
    from openpyxl import load_workbook
    hints = []
    for item in citations[:12]:
        source = str(item.get('source_path') or '')
        identifier = re.search(r'\b(?:SIG|KPI|EV|INIT|RD|CT)-[A-Za-z0-9-]+\b', str(item.get('locator') or ''))
        if not source.endswith('.xlsx') or not identifier:
            continue
        root = bundle.evidence.dataset_root.resolve(); path = (root / source).resolve()
        entry = bundle.evidence.manifest.get(source) or {}
        if not path.is_relative_to(root) or not path.is_file() or hashlib.sha256(path.read_bytes()).hexdigest() != entry.get('sha256'):
            continue
        candidates = []; book = load_workbook(path, read_only=True, data_only=True)
        try:
            for sheet in book:
                rows = sheet.iter_rows(values_only=True); headers = next(rows, ())
                columns = [i for i, key in enumerate(headers) if re.sub(r'[^a-z0-9]', '', str(key).lower()).endswith('id')]
                for number, values in enumerate(rows, 2):
                    if number > 10000:
                        break
                    if any(i < len(values) and str(values[i]).strip() == identifier[0] for i in columns):
                        candidates.append({'locator': f'{sheet.title}!Excel row {number}',
                            'row': dict(zip(map(str, headers), map(str, values)))})
            if candidates:
                hints.append({'source_path': source, 'candidate_count': len(candidates),
                    'candidates': guard_untrusted_document_text(json.dumps(candidates[-8:]),
                        source_name=source, max_chars=10000)['guarded_text']})
        finally:
            book.close()
    return hints


def verify_source_citations(bundle: DataBundle, citations: list[dict[str, Any]]) -> tuple[list[dict[str, Any]], list[str]]:
    verified, errors = [], []
    for item in citations:
        source = str(item.get('source_path') or ''); locator = str(item.get('locator') or '')
        entry = bundle.evidence.manifest.get(source)
        if source.endswith('.xlsx'):
            locator = canonical_workbook_locator(bundle, source, locator) or locator
        expected_hash = (entry or {}).get('sha256')
        if not expected_hash or item.get('source_hash') not in (None, '', expected_hash):
            errors.append(f'{source}: source hash is absent or inconsistent with the approved manifest')
            continue
        root = bundle.evidence.dataset_root.resolve(); path = (root / source).resolve()
        if not path.is_relative_to(root) or not path.is_file() or hashlib.sha256(path.read_bytes()).hexdigest() != expected_hash:
            errors.append(f'{source}: source bytes do not match the approved manifest')
            continue
        citation = Citation(source, locator, str(item.get('excerpt') or ''), source_hash=expected_hash)
        result = resolve_citation(bundle, citation)
        if not result['resolved']:
            errors.append(f'{source} / {locator}: exact source location does not resolve')
            continue
        verified.append({**item, 'locator': locator, 'source_hash': expected_hash, 'resolved': True})
    return verified, errors


def resolve_citation(bundle: DataBundle, citation: Citation) -> dict[str, Any]:
    rel_path = citation.source_path
    manifest_entry = bundle.evidence.manifest.get(rel_path, {})
    payload = resolve_payload(bundle, rel_path, citation.locator, citation.excerpt)
    source_hash = manifest_entry.get("sha256")
    validation = build_validation(rel_path, manifest_entry, source_hash, citation.source_hash, citation.locator, payload)
    return {
        "source_path": rel_path,
        "locator": citation.locator,
        "source_hash": source_hash,
        "citation_hash": citation.source_hash,
        "hash_match": validation["hash_match"],
        "manifest_entry": manifest_entry,
        "excerpt": citation.excerpt,
        "resolved_payload": payload,
        "validation": validation,
        "resolved": validation["resolved"],
    }


def resolve_findings(bundle: DataBundle, findings: list[Finding]) -> list[dict[str, Any]]:
    resolved: list[dict[str, Any]] = []
    for finding in findings:
        for citation in finding.citations:
            item = resolve_citation(bundle, citation)
            item["finding_id"] = finding.finding_id
            item["pattern_type"] = finding.pattern_type
            resolved.append(item)
    return resolved


def save_citation_audit(bundle: DataBundle, findings: list[Finding], output_path: Path) -> Path:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    records = resolve_findings(bundle, findings)
    finding_validations = validate_quantitative_claims(bundle, findings)
    payload = {
        "summary": {
            "citation_count": len(records),
            "resolved_count": sum(1 for record in records if record["resolved"]),
            "hash_match_count": sum(1 for record in records if record["hash_match"]),
            "quantitative_claim_count": len([item for item in finding_validations if item["status"] != "not_applicable"]),
            "quantitative_claim_pass_count": sum(1 for item in finding_validations if item["status"] == "pass"),
        },
        "finding_validations": finding_validations,
        "records": records,
    }
    output_path.write_text(json.dumps(payload, indent=2, default=str), encoding="utf-8")
    return output_path


def build_validation(
    rel_path: str,
    manifest_entry: dict[str, Any],
    source_hash: str | None,
    citation_hash: str | None,
    locator: str,
    payload: dict[str, Any] | None,
) -> dict[str, Any]:
    locator_type = expected_locator_type(rel_path)
    if rel_path.endswith('.xlsx') and re.fullmatch(r'.+!Excel row \d+', locator):
        locator_type = 'row'
    if rel_path.endswith('.xlsx') and re.fullmatch(r'.+!Excel rows \d+-\d+', locator):
        locator_type = 'rows'
    locator_value = parse_locator_value(locator_type, locator)
    locator_resolved = payload is not None and locator_matches_payload(locator_type, locator_value, payload)
    file_present = bool(manifest_entry)
    hash_match = bool(source_hash and citation_hash == source_hash)
    resolved = file_present and hash_match and locator_resolved
    return {
        "file_present": file_present,
        "hash_match": hash_match,
        "locator_type": locator_type,
        "locator_value": locator_value,
        "locator_resolved": locator_resolved,
        "resolved": resolved,
    }


def resolve_payload(bundle: DataBundle, rel_path: str, locator: str, excerpt: str) -> dict[str, Any] | None:
    if rel_path.endswith(".xlsx"):
        if re.fullmatch(r'.+!Excel rows? \d+(?:-\d+)?', locator):
            return resolve_manifest_workbook_row(bundle, rel_path, locator)
        if rel_path == "07_Cash_Forecast/CFO_Cash_Forecast_June_2026.xlsx":
            return resolve_workbook_sheet(bundle, locator)
        table = table_for_source(bundle, rel_path)
        if table is not None:
            return resolve_structured_row(table, locator)
        return resolve_manifest_workbook_row(bundle, rel_path, locator)
    if rel_path.endswith(".csv"):
        return resolve_structured_row(table_for_source(bundle, rel_path), locator)
    if rel_path.endswith(".pdf"):
        return resolve_pdf_page(bundle, rel_path, locator, excerpt)
    if rel_path.endswith(".json"):
        return resolve_json_record(bundle, rel_path, locator)
    if rel_path.endswith(('.txt', '.md', '.docx')) and re.fullmatch(r'(?:Text chunk|Document paragraph) \d+', locator):
        return resolve_source_text_location(bundle, rel_path, locator)
    if rel_path.endswith(".txt"):
        return resolve_text_file(bundle, rel_path)
    return {"source_type": "file", "excerpt": excerpt} if excerpt else None


def resolve_json_record(bundle: DataBundle, rel_path: str, locator: str) -> dict[str, Any] | None:
    """Resolve a unique explicit record ID, never a model-supplied excerpt."""
    if not re.fullmatch(r'[A-Za-z0-9][A-Za-z0-9_.:-]{0,159}', locator):
        return None
    root = bundle.evidence.dataset_root.resolve()
    path = (root / rel_path).resolve()
    entry = bundle.evidence.manifest.get(rel_path) or {}
    if not path.is_relative_to(root) or not path.is_file() or path.stat().st_size > 5_000_000:
        return None
    raw = path.read_bytes()
    if hashlib.sha256(raw).hexdigest() != entry.get('sha256'):
        return None
    try:
        pending = [(json.loads(raw), 0)]
    except (ValueError, RecursionError):
        return None
    found = []; visited = 0
    while pending:
        value, depth = pending.pop(); visited += 1
        if visited > 10000 or depth > 30:
            return None
        if isinstance(value, dict):
            if any((key == 'id' or key.endswith('_id')) and item == locator for key, item in value.items()):
                found.append(value)
                if len(found) > 1:
                    return None
            pending.extend((item, depth + 1) for item in value.values() if isinstance(item, (dict, list)))
        elif isinstance(value, list):
            pending.extend((item, depth + 1) for item in value if isinstance(item, (dict, list)))
    if len(found) != 1:
        return None
    return {'source_type': 'json_record', 'locator_type': 'file', 'locator_value': locator,
            'record_id': locator, 'text_excerpt': guard_untrusted_document_text(
                json.dumps(found[0], ensure_ascii=False), source_name=f'{rel_path} record {locator}',
                max_chars=12000)['guarded_text']}


def resolve_source_text_location(bundle: DataBundle, rel_path: str, locator: str) -> dict[str, Any] | None:
    from .source_search import source_records
    from types import SimpleNamespace
    entry = bundle.evidence.manifest.get(rel_path)
    if not entry:
        return None
    evidence = SimpleNamespace(dataset_root=bundle.evidence.dataset_root, manifest={rel_path:entry}, pdf_text={})
    chunks = [text for _, _, location, text in source_records(evidence) if location == locator]
    if not chunks:
        return None
    return {'source_type':'text', 'locator_type':'file', 'locator_value':locator,
            'text_excerpt':guard_untrusted_document_text(' '.join(chunks), source_name=rel_path, max_chars=2000)['guarded_text']}


def resolve_manifest_workbook_row(bundle: DataBundle, rel_path: str, locator: str) -> dict[str, Any] | None:
    """Resolve a source-addressed workbook row outside the legacy table map."""
    match = re.fullmatch(r"(.+)!Excel rows? (\d+)(?:-(\d+))?", locator)
    entry = bundle.evidence.manifest.get(rel_path)
    root = bundle.evidence.dataset_root.resolve()
    path = (root / rel_path).resolve()
    if not match or not entry or not path.is_relative_to(root) or not path.is_file():
        return None
    if hashlib.sha256(path.read_bytes()).hexdigest() != entry.get("sha256"):
        return None
    from openpyxl import load_workbook
    book = load_workbook(path, read_only=True, data_only=True)
    try:
        if match[1] not in book.sheetnames:
            return None
        sheet = book[match[1]]; row_number = int(match[2]); end = int(match[3] or match[2])
        if row_number < 2 or end < row_number or end > 10000 or (sheet.max_row is not None and end > sheet.max_row):
            return None
        headers = next(sheet.iter_rows(min_row=1,max_row=1,values_only=True))
        if match[3]:
            rows = list(sheet.iter_rows(min_row=row_number,max_row=end,values_only=True))
            if len(rows) != end-row_number+1 or any(all(value is None for value in row) for row in rows):
                return None
            return {"source_type":"structured_table", "locator_type":"rows", "locator_value":f'{row_number}-{end}',
                    "sheet_name":sheet.title, "rows":[{str(key):normalize_value(value) for key,value in zip(headers,row) if key is not None} for row in rows]}
        values = next(sheet.iter_rows(min_row=row_number,max_row=row_number,values_only=True), None)
        if values is None:
            return None
        return {"source_type": "structured_table", "locator_type": "row", "locator_value": row_number,
                "sheet_name": sheet.title, "row": {str(key): normalize_value(value) for key,value in zip(headers,values) if key is not None}}
    finally:
        book.close()


def resolve_structured_row(df: pd.DataFrame | None, locator: str) -> dict[str, Any] | None:
    if df is None:
        return None
    match = ROW_RE.search(locator)
    if not match:
        return {"source_type": "structured_table", "row_count": int(len(df)), "columns": list(df.columns)}
    row_index = int(match.group(1)) - 2
    if row_index < 0 or row_index >= len(df):
        return None
    row = df.iloc[row_index]
    return {
        "source_type": "structured_table",
        "locator_type": "row",
        "locator_value": row_index + 2,
        "row_index_zero_based": row_index,
        "columns": list(df.columns),
        "row": {str(k): normalize_value(v) for k, v in row.to_dict().items()},
    }


def resolve_workbook_sheet(bundle: DataBundle, locator: str) -> dict[str, Any] | None:
    sheet_name = locator.replace(" sheet", "").strip()
    df = bundle.cash_forecast.get(sheet_name)
    if df is None:
        return None
    records = [
        {str(k): normalize_value(v) for k, v in row.items()}
        for row in df.head(10).to_dict(orient="records")
    ]
    return {
        "source_type": "workbook_sheet",
        "locator_type": "sheet",
        "locator_value": sheet_name,
        "sheet_name": sheet_name,
        "row_count": int(len(df)),
        "columns": list(df.columns),
        "sample_records": records,
    }


def resolve_pdf_page(bundle: DataBundle, rel_path: str, locator: str, excerpt: str) -> dict[str, Any] | None:
    pages = bundle.evidence.pdf_text.get(rel_path, [])
    page_no = page_number(locator) or page_number(excerpt)
    if page_no is None:
        return None
    if page_no < 1 or page_no > len(pages):
        return None
    text = " ".join(pages[page_no - 1].split())
    if not text:
        return None
    return {
        "source_type": "pdf",
        "locator_type": "page",
        "locator_value": page_no,
        "page": page_no,
        "text_excerpt": guard_untrusted_document_text(
            text[:700],
            source_name=f"{rel_path} page {page_no}",
            max_chars=700,
        )["guarded_text"],
    }


def resolve_text_file(bundle: DataBundle, rel_path: str) -> dict[str, Any] | None:
    path = bundle.dataset_root / rel_path
    if not path.exists():
        return None
    text = " ".join(path.read_text(encoding="utf-8", errors="ignore").split())
    if not text:
        return None
    return {
        "source_type": "text",
        "locator_type": "file",
        "locator_value": rel_path,
        "text_excerpt": guard_untrusted_document_text(
            text[:700],
            source_name=rel_path,
            max_chars=700,
        )["guarded_text"],
    }


def table_for_source(bundle: DataBundle, rel_path: str) -> pd.DataFrame | None:
    mapping = {
        "02_ERP_Extracts/AP_Invoices_H1_2026.xlsx": bundle.ap,
        "02_ERP_Extracts/AR_Invoices_H1_2026.xlsx": bundle.ar,
        "02_ERP_Extracts/GL_Extract_H1_2026.csv": bundle.gl,
        "02_ERP_Extracts/Trial_Balance_June_2026.xlsx": bundle.trial_balance,
        "03_Master_Data/Vendor_Master.xlsx": bundle.vendors,
        "03_Master_Data/Customer_Master.xlsx": bundle.customers,
        "03_Master_Data/Chart_of_Accounts.xlsx": bundle.coa,
        "05_Purchase_Orders/PO_Log_H1_2026.csv": bundle.po,
    }
    return mapping.get(rel_path)


def page_number(text: str) -> int | None:
    match = PAGE_RE.search(text or "")
    return int(match.group(1)) if match else None


def expected_locator_type(rel_path: str) -> str:
    if rel_path == "07_Cash_Forecast/CFO_Cash_Forecast_June_2026.xlsx":
        return "sheet"
    if rel_path.endswith((".xlsx", ".csv")):
        return "row"
    if rel_path.endswith(".pdf"):
        return "page"
    return "file"


def parse_locator_value(locator_type: str, locator: str) -> Any:
    if locator_type == 'rows':
        match = re.fullmatch(r'.+!Excel rows (\d+-\d+)', locator)
        return match[1] if match else None
    if locator_type == "row":
        match = ROW_RE.search(locator or "")
        return int(match.group(1)) if match else None
    if locator_type == "page":
        return page_number(locator)
    if locator_type == "sheet":
        return locator.replace(" sheet", "").strip() if locator else None
    return locator or None


def locator_matches_payload(locator_type: str, locator_value: Any, payload: dict[str, Any]) -> bool:
    if not locator_value:
        return False
    if locator_type in {"row", "rows"}:
        return payload.get("locator_type") == locator_type and payload.get("locator_value") == locator_value
    if locator_type == "page":
        return payload.get("locator_type") == "page" and payload.get("locator_value") == locator_value
    if locator_type == "sheet":
        return payload.get("locator_type") == "sheet" and payload.get("locator_value") == locator_value
    return payload.get("locator_type") == "file"


def validate_quantitative_claims(bundle: DataBundle, findings: list[Finding]) -> list[dict[str, Any]]:
    return [validate_quantitative_claim(bundle, finding) for finding in findings]


def validate_quantitative_claim(bundle: DataBundle, finding: Finding) -> dict[str, Any]:
    if finding.pattern_type == "duplicate_payment":
        return _validate_duplicate_payment(bundle, finding)
    if finding.pattern_type == "entity_resolution_duplicate":
        return _validate_entity_resolution_duplicate(bundle, finding)
    if finding.pattern_type == "price_variance":
        return _validate_price_variance(bundle, finding)
    return {
        "finding_id": finding.finding_id,
        "pattern_type": finding.pattern_type,
        "status": "not_applicable",
        "checks": [],
    }


def _validate_duplicate_payment(bundle: DataBundle, finding: Finding) -> dict[str, Any]:
    ap_rows = _rows_for_source(bundle, finding, "02_ERP_Extracts/AP_Invoices_H1_2026.xlsx")
    amount = float(finding.calculation.get("amount_sar", 0.0))
    invoice_ids = {str(row.get("Invoice_ID")) for row in ap_rows}
    ap_ok = len(ap_rows) >= 2 and len(invoice_ids) == 1 and all(_float_value(row.get("Amount_SAR")) == amount for row in ap_rows)
    bank_records = [
        resolve_citation(bundle, citation)
        for citation in finding.citations
        if citation.source_path.startswith("01_Bank_Statements/")
    ]
    bank_ok = len(bank_records) >= 2 and all(record["resolved"] for record in bank_records)
    checks = [
        {"name": "ap_rows", "passed": ap_ok, "detail": f"resolved_ap_rows={len(ap_rows)} invoice_ids={sorted(invoice_ids)} amount={amount:,.2f}"},
        {"name": "bank_statement_payment_legs", "passed": bank_ok, "detail": f"resolved_bank_legs={sum(1 for record in bank_records if record['resolved'])}"},
    ]
    return _claim_result(finding, checks)


def _validate_entity_resolution_duplicate(bundle: DataBundle, finding: Finding) -> dict[str, Any]:
    vendor_rows = _rows_for_source(bundle, finding, "03_Master_Data/Vendor_Master.xlsx")
    ap_rows = _rows_for_source(bundle, finding, "02_ERP_Extracts/AP_Invoices_H1_2026.xlsx")
    vendor_ids = set(str(v) for v in finding.vendor_id.split("/"))
    shared_field = str(finding.calculation.get("identity_field", ""))
    shared_value = str(finding.calculation.get("identity_value", ""))
    vendor_ok = {str(row.get("Vendor_ID")) for row in vendor_rows} == vendor_ids and all(
        matches_sensitive_identifier_token(shared_value, row.get(shared_field), field_name=shared_field) for row in vendor_rows
    )
    ap_ok = vendor_ids.issubset({str(row.get("Vendor_ID")) for row in ap_rows})
    checks = [
        {"name": "vendor_master_rows", "passed": vendor_ok, "detail": f"resolved_vendor_ids={sorted(str(row.get('Vendor_ID')) for row in vendor_rows)} shared={shared_field}={shared_value}"},
        {"name": "ap_vendor_coverage", "passed": ap_ok, "detail": f"ap_vendor_ids={sorted({str(row.get('Vendor_ID')) for row in ap_rows})}"},
    ]
    return _claim_result(finding, checks)


def _validate_price_variance(bundle: DataBundle, finding: Finding) -> dict[str, Any]:
    po_rows = _rows_for_source(bundle, finding, "05_Purchase_Orders/PO_Log_H1_2026.csv")
    ap_rows = _rows_for_source(bundle, finding, "02_ERP_Extracts/AP_Invoices_H1_2026.xlsx")
    baseline = float(finding.calculation.get("baseline_unit_price", 0.0))
    high = float(finding.calculation.get("high_unit_price", 0.0))
    po_ok = (
        len(po_rows) >= 2
        and any(_float_value(row.get("Unit_Price")) == baseline for row in po_rows)
        and any(_float_value(row.get("Unit_Price")) == high for row in po_rows)
        and all(str(row.get("Vendor_ID")) == finding.vendor_id for row in po_rows)
    )
    po_ids = {str(row.get("PO_ID")) for row in po_rows}
    ap_ok = len(ap_rows) >= 2 and po_ids.issubset({str(row.get("PO_Reference")) for row in ap_rows}) and all(str(row.get("Vendor_ID")) == finding.vendor_id for row in ap_rows)
    contract_records = [
        resolve_citation(bundle, citation)
        for citation in finding.citations
        if citation.source_path.startswith("04_Contracts/")
    ]
    contract_ok = any(record["resolved"] and str(po_rows[0].get("SKU")) in record["excerpt"] and f"{baseline:.2f}" in record["excerpt"] for record in contract_records if po_rows)
    checks = [
        {"name": "po_rows", "passed": po_ok, "detail": f"po_ids={sorted(po_ids)} baseline={baseline:.2f} high={high:.2f}"},
        {"name": "matching_ap_rows", "passed": ap_ok, "detail": f"ap_po_refs={sorted({str(row.get('PO_Reference')) for row in ap_rows})}"},
        {"name": "contract_price_schedule", "passed": contract_ok, "detail": f"contract_citations={len(contract_records)} baseline={baseline:.2f}"},
    ]
    return _claim_result(finding, checks)


def _rows_for_source(bundle: DataBundle, finding: Finding, source_path: str) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for citation in finding.citations:
        if citation.source_path != source_path:
            continue
        resolved = resolve_citation(bundle, citation)
        payload = resolved.get("resolved_payload") or {}
        row = payload.get("row")
        if resolved["resolved"] and isinstance(row, dict):
            rows.append(row)
    return rows


def _claim_result(finding: Finding, checks: list[dict[str, Any]]) -> dict[str, Any]:
    return {
        "finding_id": finding.finding_id,
        "pattern_type": finding.pattern_type,
        "status": "pass" if all(check["passed"] for check in checks) else "fail",
        "checks": checks,
    }


def _float_value(value: Any) -> float:
    if value is None:
        return 0.0
    return round(float(value), 2)


def normalize_value(value: Any) -> Any:
    if isinstance(value, pd.Timestamp):
        return value.isoformat()
    if pd.isna(value):
        return None
    if hasattr(value, "item"):
        return value.item()
    return value
