"""Deterministic CEO finance KPIs from the files supplied with a run.

This adapter intentionally does not pretend that an uploaded spreadsheet is an
Oracle snapshot.  It derives only the four CEO actuals that can be reproduced
from the source pack, and keeps the selected files, accounts and row coverage
with the result so the presentation and Hermes can explain every number.
"""

from __future__ import annotations

import csv
import hashlib
from collections import defaultdict
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any, Iterable, Mapping

from openpyxl import load_workbook


def derive_source_finance_kpis(dataset_root: Path) -> dict[str, Any]:
    """Return a JSON-safe calculation payload, or an explicit unavailable payload.

    The calculation boundary is deliberately narrow: GL balances calculate
    Revenue, EBITDA and operating cost; the CFO cash-position extract supplies
    cash. Plans and the board floor are never manufactured from actuals.
    """
    root = Path(dataset_root)
    # A group finance pack has a higher-quality, explicitly aligned CEO view
    # than a division GL extract.  Prefer it when all of its required source
    # workbooks are present; otherwise retain the existing narrow GL path.
    group_projection = _group_finance_projection(root)
    if group_projection is not None:
        return group_projection
    gl_path = _first_matching(root, "gl_extract", ".csv")
    coa_path = _first_matching(root, "chart_of_accounts", ".xlsx")
    cash_path = _first_matching(root, "cash_forecast", ".xlsx")
    if gl_path is None or coa_path is None:
        return _unavailable(
            "A GL extract and chart of accounts are required to derive the CEO finance actuals.",
            root,
        )

    accounts = _load_accounts(coa_path)
    gl = _load_gl(gl_path)
    if not gl:
        return _unavailable("The detected GL extract contains no usable debit/credit rows.", root)

    balances: dict[str, Decimal] = defaultdict(Decimal)
    dates: list[date] = []
    row_count = 0
    for row in gl:
        account = str(row.get("account") or "").strip()
        if not account:
            continue
        debit = _decimal(row.get("debit"))
        credit = _decimal(row.get("credit"))
        if debit is None or credit is None:
            continue
        balances[account] += debit - credit
        row_count += 1
        parsed_date = _date_value(row.get("date"))
        if parsed_date is not None:
            dates.append(parsed_date)

    revenue_accounts = _accounts_matching(balances, accounts, _is_revenue)
    cogs_accounts = _accounts_matching(balances, accounts, _is_cogs)
    operating_cost_accounts = _accounts_matching(balances, accounts, _is_operating_cost)
    if not revenue_accounts or not cogs_accounts or not operating_cost_accounts:
        return _unavailable(
            "The GL/Chart of Accounts mapping does not contain the revenue, COGS and operating-expense scopes required for EBITDA.",
            root,
        )

    revenue = sum((-balances[account] for account in revenue_accounts), Decimal())
    cogs = sum((balances[account] for account in cogs_accounts), Decimal())
    operating_cost = sum((balances[account] for account in operating_cost_accounts), Decimal())
    ebitda = revenue - cogs - operating_cost
    period = _period_label(dates)
    # Derive the like-for-like plan before the trend so the revenue movement can
    # carry a phased plan line when an aligned budget exists.
    plan = _reconciliation_plan(root, period)
    revenue_plan = plan.get("revenue_plan") if plan else None
    plan_basis = plan.get("basis") if plan else None
    revenue_plan_total = _decimal_or_none(revenue_plan)
    finance_dynamics = _finance_dynamics(
        gl,
        revenue_accounts,
        cogs_accounts,
        operating_cost_accounts,
        accounts,
        revenue_plan_total=revenue_plan_total,
    )
    gl_evidence = {
        "file": _relative(gl_path, root),
        "sha256": _sha256(gl_path),
        "usable_rows": row_count,
        "period_start": min(dates).isoformat() if dates else None,
        "period_end": max(dates).isoformat() if dates else None,
        "chart_of_accounts_file": _relative(coa_path, root),
        "chart_of_accounts_sha256": _sha256(coa_path),
        "account_scopes": {
            "revenue": _scope(revenue_accounts, balances),
            "cogs": _scope(cogs_accounts, balances),
            "operating_cost": _scope(operating_cost_accounts, balances),
            "excluded_from_ebitda": {
                "rule": "Depreciation, amortization and interest are excluded from EBITDA.",
                "accounts": [
                    account for account in sorted(balances, key=_account_sort_key)
                    if _is_ebitda_exclusion(account, accounts.get(account, {}))
                ],
            },
        },
        "contributors": {
            "revenue": _contributors(revenue_accounts, balances, accounts, credit_nature=True),
            "cogs": _contributors(cogs_accounts, balances, accounts),
            "operating_cost": _contributors(operating_cost_accounts, balances, accounts),
        },
    }
    cash = _cash_position(cash_path, root)
    cash_evidence = cash.pop("evidence", {})
    components: dict[str, str | None] = {
        "revenue_actual": _number(revenue),
        "revenue_plan": revenue_plan,
        "cogs_actual": _number(cogs),
        "ebitda_actual": _number(ebitda),
        "ebitda_plan": None,
        "operating_cost_actual": _number(operating_cost),
        "operating_cost_plan": None,
        "cash_balance": cash.get("value"),
        "board_floor": None,
    }
    actual_complete = {
        "revenue": True,
        "ebitda_margin": True,
        "operating_cost": True,
        "cash_vs_floor": bool(cash.get("complete")),
    }
    evidence = {
        "revenue": {
            "summary": f"{row_count:,} GL rows across {len(revenue_accounts)} scoped revenue accounts.",
            "files": (
                [gl_evidence["file"], gl_evidence["chart_of_accounts_file"], plan_basis["source_file"]]
                if plan_basis
                else [gl_evidence["file"], gl_evidence["chart_of_accounts_file"]]
            ),
            "actual_complete": True,
            "details": gl_evidence,
            # When a like-for-like plan was derived, its full provenance rides
            # with the revenue KPI so the comparison can be defended.
            "plan_basis": plan_basis,
        },
        "ebitda_margin": {
            "summary": (
                f"Revenue ({_number(revenue)} SAR) less COGS ({_number(cogs)} SAR) "
                f"and operating cost ({_number(operating_cost)} SAR); depreciation, amortization and interest excluded."
            ),
            "files": [gl_evidence["file"], gl_evidence["chart_of_accounts_file"]],
            "actual_complete": True,
            "details": gl_evidence,
        },
        "operating_cost": {
            "summary": f"{len(operating_cost_accounts)} scoped operating-expense accounts, excluding depreciation, amortization and interest.",
            "files": [gl_evidence["file"], gl_evidence["chart_of_accounts_file"]],
            "actual_complete": True,
            "details": gl_evidence,
        },
        "cash_vs_floor": {
            "summary": str(cash.get("summary") or "No usable cash-position extract was found."),
            "files": list(cash_evidence.get("files") or []),
            "actual_complete": bool(cash.get("complete")),
            "details": cash_evidence,
        },
    }
    return {
        "authoritative": True,
        "derived_from": "deterministic_source_finance_kpi_engine",
        "reporting_period_key": period,
        "reporting_currency": "SAR",
        "computation_boundary": (
            "Actuals are calculated only from the listed uploaded source extracts. "
            "No plan or board floor is inferred when it is absent."
        ),
        "components": components,
        # Revenue movement is a period-by-period calculation from the same GL
        # rows as the headline actual.  A plan is deliberately absent unless a
        # separately governed budget role supplies one; actuals never become a
        # proxy plan.
        "trend": finance_dynamics["trend"],
        "dynamics": finance_dynamics["movers"],
        "actual_complete": actual_complete,
        "evidence": evidence,
        "source_files": sorted({item for group in evidence.values() for item in group["files"]}),
    }


def _group_finance_projection(root: Path) -> dict[str, Any] | None:
    """Derive CEO KPI cards from the supplied group P&L, budget and analytics.

    The source pack contains a governed H1 group budget by business unit and a
    separately documented monthly revenue series for the Tamween division.
    These must not be silently mixed: headline group actuals and plan values
    come only from ``BU_Group_Budget_2026``; the trend is explicitly labelled
    as division steering data.  Each mover is the workbook's stated H1
    variance and keeps the source note as an expandable source note.
    """
    budget_path = _first_matching(root, "bu_group_budget_2026", ".xlsx")
    analytics_path = _first_matching(root, "revenue_analytics_2023-2026", ".xlsx")
    if budget_path is None:
        return None
    try:
        budget_book = load_workbook(budget_path, data_only=True, read_only=True)
        budget_sheet = budget_book["BU_Budget_2026"]
        rows = list(budget_sheet.values)
    except Exception:
        return None
    if len(rows) < 2:
        return None

    headers = _header_positions(rows[0])
    required = ("businessunit", "h1budget", "h1actualestsarm", "h1var", "ebitdabudget", "ebitdah1est")
    if any(key not in headers for key in required):
        return None
    units: list[dict[str, Any]] = []
    group_total: dict[str, Decimal] | None = None
    for values in rows[1:]:
        unit = str(_cell(values, headers, "businessunit") or "").strip()
        actual = _decimal(_cell(values, headers, "h1actualestsarm"))
        plan = _decimal(_cell(values, headers, "h1budget"))
        variance = _decimal(_cell(values, headers, "h1var"))
        plan_margin = _decimal(_cell(values, headers, "ebitdabudget"))
        actual_margin = _decimal(_cell(values, headers, "ebitdah1est"))
        if _normal(unit) == "group" and None not in {actual, plan, plan_margin, actual_margin}:
            group_total = {"actual": actual, "plan": plan, "actual_margin": actual_margin, "plan_margin": plan_margin}
            continue
        # The GROUP total above is used for CEO headlines, but is never a
        # business-unit mover.  Rows without a stated EBITDA margin (such as
        # eliminations) are also not suitable for a margin reconstruction.
        if not unit or None in {actual, plan, variance, plan_margin, actual_margin}:
            continue
        units.append({
            "name": unit,
            "actual": actual,
            "plan": plan,
            "variance": variance,
            "plan_margin": plan_margin,
            "actual_margin": actual_margin,
            "note": str(_cell(values, headers, "note") or "").strip(),
        })
    if not units:
        return None

    # Workbook financial amounts are SAR millions.  Convert only at the
    # calculation boundary so the API remains consistently in SAR.
    million = Decimal("1000000")
    revenue_actual = (group_total["actual"] if group_total else sum((row["actual"] for row in units), Decimal())) * million
    revenue_plan = (group_total["plan"] if group_total else sum((row["plan"] for row in units), Decimal())) * million
    ebitda_actual = ((revenue_actual / million) * group_total["actual_margin"] / 100 if group_total else sum((row["actual"] * row["actual_margin"] / 100 for row in units), Decimal())) * million
    ebitda_plan = ((revenue_plan / million) * group_total["plan_margin"] / 100 if group_total else sum((row["plan"] * row["plan_margin"] / 100 for row in units), Decimal())) * million
    operating_cost_actual = revenue_actual - ebitda_actual
    operating_cost_plan = revenue_plan - ebitda_plan
    cash = _group_cash_floor(budget_book, budget_path, root)
    trend = _group_revenue_trend(analytics_path, root)
    # The group pack supplies one consolidated H1 headline, while the same
    # governed source set also carries monthly Tamween division actuals and a
    # quarterly group cash series. Use those explicitly labelled series for
    # trajectory charts; never manufacture a history from the headline.
    gl_path = _first_matching(root, "gl_extract_h1_2026", ".csv")
    coa_path = _first_matching(root, "chart_of_accounts", ".xlsx")
    if gl_path is not None and coa_path is not None:
        accounts = _load_accounts(coa_path)
        gl = _load_gl(gl_path)
        balances: dict[str, Decimal] = defaultdict(Decimal)
        for row in gl:
            account = str(row.get("account") or "").strip()
            debit = _decimal(row.get("debit"))
            credit = _decimal(row.get("credit"))
            if account and debit is not None and credit is not None:
                balances[account] += debit - credit
        revenue_accounts = _accounts_matching(balances, accounts, _is_revenue)
        cogs_accounts = _accounts_matching(balances, accounts, _is_cogs)
        operating_accounts = _accounts_matching(balances, accounts, _is_operating_cost)
        if revenue_accounts and cogs_accounts and operating_accounts:
            division_dynamics = _finance_dynamics(
                gl,
                revenue_accounts,
                cogs_accounts,
                operating_accounts,
                accounts,
            )["trend"]
            trend["ebitda_margin"] = {
                **division_dynamics["ebitda_margin"],
                "scope_note": "Tamween division monthly ledger actuals; group H1 headline is shown above.",
            }
            trend["operating_cost"] = {
                **division_dynamics["total_cost_to_ebitda"],
                "scope_note": "Tamween division monthly COGS plus cash operating cost; group H1 headline is shown above.",
            }
            trend["source_files"] = sorted({
                *trend.get("source_files", []),
                _relative(gl_path, root),
                _relative(coa_path, root),
            })
    trend["cash_vs_floor"] = _group_cash_floor_trend(budget_book)
    movers = _group_movers(units)
    budget_file = _relative(budget_path, root)
    budget_sha = _sha256(budget_path)
    period = "H1 2026"
    evidence_base = {
        "files": [budget_file],
        "actual_complete": True,
        "details": {"file": budget_file, "sha256": budget_sha, "sheet": "BU_Budget_2026", "unit_count": len(units)},
    }
    evidence = {
        "revenue": {**evidence_base, "summary": f"H1 actual and plan aggregated across {len(units)} business units from the approved group budget."},
        "ebitda_margin": {**evidence_base, "summary": "H1 EBITDA is reconstructed from each business unit's H1 revenue and stated EBITDA margin; no group allocation has been added."},
        "operating_cost": {**evidence_base, "summary": "Total cost to EBITDA is revenue less the stated business-unit EBITDA, not a proxy for a separately supplied opex ledger."},
        "cash_vs_floor": cash["evidence"],
    }
    return {
        "authoritative": True,
        "derived_from": "deterministic_source_finance_kpi_engine",
        "reporting_period_key": period,
        "reporting_currency": "SAR",
        "computation_boundary": "Group headlines use BU_Group_Budget_2026 H1 values. The revenue trend is separately labelled Tamween division steering data. No unprovided group cost allocation is inferred.",
        "components": {
            "revenue_actual": _number(revenue_actual), "revenue_plan": _number(revenue_plan),
            "ebitda_actual": _number(ebitda_actual), "ebitda_plan": _number(ebitda_plan),
            "operating_cost_actual": _number(operating_cost_actual), "operating_cost_plan": _number(operating_cost_plan),
            "cash_balance": cash["value"], "board_floor": cash["floor"],
        },
        "trend": trend,
        "dynamics": movers,
        "actual_complete": {"revenue": True, "ebitda_margin": True, "operating_cost": True, "cash_vs_floor": cash["complete"]},
        "evidence": evidence,
        "source_files": sorted({*([budget_file]), *trend.get("source_files", []), *cash["evidence"].get("files", [])}),
    }


def _header_positions(values: Iterable[Any]) -> dict[str, int]:
    return {_normal(value): index for index, value in enumerate(values) if value is not None}


def _cell(values: Iterable[Any], headers: Mapping[str, int], key: str) -> Any:
    items = tuple(values)
    index = headers.get(key)
    return items[index] if index is not None and index < len(items) else None


def _group_movers(units: list[dict[str, Any]]) -> dict[str, list[dict[str, str]]]:
    lifting: list[dict[str, str]] = []
    dragging: list[dict[str, str]] = []
    for unit in sorted(units, key=lambda item: item["variance"], reverse=True):
        variance = unit["variance"]
        if variance == 0:
            continue
        row = {"name": str(unit["name"]), "delta": _sar_delta(variance * Decimal("1000000"))}
        if unit["note"]:
            row["gm"] = "BU note"
            row["note"] = unit["note"]
        (lifting if variance > 0 else dragging).append(row)
    return {"revenue": {"lifting": lifting[:4], "dragging": dragging[:4]}}


def _group_revenue_trend(path: Path | None, root: Path) -> dict[str, Any]:
    empty = {"labels": [], "actual": [], "plan": [], "has_plan_series": False, "unit": "sar"}
    result = {"revenue": empty, "ebitda_margin": empty, "operating_cost": empty, "source_files": []}
    if path is None:
        return result
    try:
        sheet = load_workbook(path, data_only=True, read_only=True)["Monthly_by_Segment"]
        rows = list(sheet.values)
    except Exception:
        return result
    headers = _header_positions(rows[0] if rows else ())
    if "month" not in headers or "netrevenuesar" not in headers:
        return result
    labels: list[str] = []
    actual: list[str] = []
    for values in rows[1:]:
        month = str(_cell(values, headers, "month") or "")
        value = _decimal(_cell(values, headers, "netrevenuesar"))
        if month.startswith("2026-") and value is not None:
            labels.append(month)
            actual.append(_number(value) or "0")
    if labels:
        result["revenue"] = {"labels": labels, "actual": actual, "plan": [], "has_plan_series": False, "unit": "sar", "scope_note": "Tamween division monthly steering data; group headline is shown above."}
        result["source_files"] = [_relative(path, root)]
    return result


def _group_cash_floor(book: Any, budget_path: Path, root: Path) -> dict[str, Any]:
    fallback = {"value": None, "floor": None, "complete": False, "evidence": {"files": [_relative(budget_path, root)], "summary": "No group cash-floor row is available."}}
    try:
        sheet = book["Group_Cash_Floor"]
        rows = list(sheet.values)
        headers = _header_positions(rows[0] if rows else ())
    except Exception:
        return fallback
    candidates = []
    for values in rows[1:]:
        quarter = str(_cell(values, headers, "quarter") or "")
        value = _decimal(_cell(values, headers, "actualforecastsarb"))
        floor = _decimal(_cell(values, headers, "floorsarb"))
        if "2026-q" in quarter.lower() and value is not None and floor is not None:
            candidates.append((quarter, value, floor))
    if not candidates:
        return fallback
    quarter, value, floor = candidates[-1]
    return {"value": _number(value * Decimal("1000000000")), "floor": _number(floor * Decimal("1000000000")), "complete": True, "evidence": {"files": [_relative(budget_path, root)], "summary": f"{quarter} group cash actual/forecast and approved floor from Group_Cash_Floor.", "details": {"file": _relative(budget_path, root), "sha256": _sha256(budget_path), "sheet": "Group_Cash_Floor", "quarter": quarter}}}


def _group_cash_floor_trend(book: Any) -> dict[str, Any]:
    empty = {
        "labels": [],
        "actual": [],
        "plan": [],
        "has_plan_series": False,
        "unit": "sar",
    }
    try:
        rows = list(book["Group_Cash_Floor"].values)
        headers = _header_positions(rows[0] if rows else ())
    except Exception:
        return empty
    labels: list[str] = []
    actual: list[str] = []
    floors: list[str] = []
    for values in rows[1:]:
        quarter = str(_cell(values, headers, "quarter") or "").strip()
        value = _decimal(_cell(values, headers, "actualforecastsarb"))
        floor = _decimal(_cell(values, headers, "floorsarb"))
        if not quarter or value is None or floor is None:
            continue
        labels.append(quarter.split(" (", 1)[0])
        actual.append(_number(value * Decimal("1000000000")) or "0")
        floors.append(_number(floor * Decimal("1000000000")) or "0")
    if len(actual) < 2:
        return empty
    return {
        "labels": labels,
        "actual": actual,
        "plan": floors,
        "has_plan_series": True,
        "unit": "sar",
        "scope_note": "Quarterly group cash actual/forecast versus the approved group floor.",
    }


def _finance_dynamics(
    gl: Iterable[Mapping[str, Any]],
    revenue_accounts: Iterable[str],
    cogs_accounts: Iterable[str],
    operating_cost_accounts: Iterable[str],
    account_master: Mapping[str, Mapping[str, str]],
    *,
    revenue_plan_total: Decimal | None = None,
) -> dict[str, Any]:
    """Calculate governed monthly finance movement from the GL.

    Revenue, COGS and operating expense are independently calculated for every
    month represented by the GL.  That gives the CEO an actual EBITDA-margin
    trajectory without fabricating a plan line.  A plan remains empty until an
    aligned, governed budget source is supplied.
    """
    revenue_set = set(revenue_accounts)
    cogs_set = set(cogs_accounts)
    operating_cost_set = set(operating_cost_accounts)
    scoped_accounts = revenue_set | cogs_set | operating_cost_set
    periods: dict[str, dict[str, Decimal]] = defaultdict(
        lambda: {
            "revenue": Decimal(),
            "cogs": Decimal(),
            "operating_cost": Decimal(),
        }
    )
    account_periods: dict[str, dict[str, Decimal]] = defaultdict(lambda: defaultdict(Decimal))
    for row in gl:
        account = str(row.get("account") or "").strip()
        current = _date_value(row.get("date"))
        if account not in scoped_accounts or current is None:
            continue
        debit = _decimal(row.get("debit"))
        credit = _decimal(row.get("credit"))
        if debit is None or credit is None:
            continue
        period_key = current.strftime("%Y-%m")
        if account in revenue_set:
            # Revenue is credit-natured, so credit less debit is the reported
            # positive contribution for the period.
            value = credit - debit
            periods[period_key]["revenue"] += value
            account_periods[account][period_key] += value
        elif account in cogs_set:
            periods[period_key]["cogs"] += debit - credit
        else:
            periods[period_key]["operating_cost"] += debit - credit

    labels = sorted(periods)
    revenue_actual = [periods[label]["revenue"] for label in labels]
    operating_cost_actual = [periods[label]["operating_cost"] for label in labels]
    total_cost_to_ebitda_actual = [
        periods[label]["cogs"] + periods[label]["operating_cost"] for label in labels
    ]
    ebitda_margin_labels: list[str] = []
    ebitda_margin_actual: list[Decimal] = []
    for label in labels:
        revenue = periods[label]["revenue"]
        if revenue == 0:
            continue
        ebitda = revenue - periods[label]["cogs"] - periods[label]["operating_cost"]
        ebitda_margin_labels.append(label)
        ebitda_margin_actual.append((ebitda / revenue) * Decimal("100"))

    def actual_trend(
        series_labels: list[str],
        values: list[Decimal],
        *,
        unit: str,
        plan_total: Decimal | None = None,
        plan_note: str | None = None,
    ) -> dict[str, Any]:
        # A monthly plan line is only drawn when a period plan total exists AND
        # the dataset gives no monthly plan of its own -- in which case the
        # annual figure is spread evenly across the actual months and labelled
        # as straight-line phasing, never passed off as a real monthly budget.
        plan_series: list[str] = []
        has_plan_series = False
        if plan_total is not None and series_labels:
            per_month = (plan_total / Decimal(len(series_labels))).quantize(Decimal("0.01"))
            plan_series = [_number(per_month) for _ in series_labels]
            has_plan_series = True
        payload: dict[str, Any] = {
            "labels": series_labels,
            "actual": [_number(value) for value in values],
            "plan": plan_series,
            "has_plan_series": has_plan_series,
            "unit": unit,
        }
        if plan_note:
            payload["plan_note"] = plan_note
        return payload

    movers: dict[str, list[dict[str, str]]] = {"lifting": [], "dragging": []}
    if len(labels) >= 2:
        previous, latest = labels[-2], labels[-1]
        deltas: list[tuple[str, Decimal]] = []
        for account, by_period in account_periods.items():
            delta = by_period.get(latest, Decimal()) - by_period.get(previous, Decimal())
            if delta:
                deltas.append((account, delta))
        for account, delta in sorted(deltas, key=lambda item: item[1], reverse=True)[:4]:
            if delta <= 0:
                continue
            label = str(account_master.get(account, {}).get("account_description") or f"Account {account}")
            movers["lifting"].append({"name": label, "delta": _sar_delta(delta)})
        for account, delta in sorted(deltas, key=lambda item: item[1])[:4]:
            if delta >= 0:
                continue
            label = str(account_master.get(account, {}).get("account_description") or f"Account {account}")
            movers["dragging"].append({"name": label, "delta": _sar_delta(delta)})
    return {
        "trend": {
            "revenue": actual_trend(
                labels,
                revenue_actual,
                unit="sar",
                plan_total=revenue_plan_total,
                plan_note=(
                    "Plan line is the approved period budget spread evenly across the months; "
                    "the dataset supplies an annual budget, not a monthly one."
                    if revenue_plan_total is not None
                    else None
                ),
            ),
            "ebitda_margin": actual_trend(ebitda_margin_labels, ebitda_margin_actual, unit="percent"),
            "operating_cost": actual_trend(labels, operating_cost_actual, unit="sar"),
            "total_cost_to_ebitda": actual_trend(labels, total_cost_to_ebitda_actual, unit="sar"),
        },
        "movers": {"revenue": movers},
    }


def _unavailable(reason: str, root: Path) -> dict[str, Any]:
    return {
        "authoritative": False,
        "derived_from": "deterministic_source_finance_kpi_engine",
        "reason": reason,
        "source_files": [str(path.relative_to(root)) for path in root.rglob("*") if path.is_file()][:20],
    }


def _first_matching(root: Path, name_fragment: str, suffix: str) -> Path | None:
    fragment = name_fragment.lower()
    for path in sorted(root.rglob(f"*{suffix}")):
        if fragment in path.name.lower():
            return path
    return None


def _load_accounts(path: Path) -> dict[str, dict[str, str]]:
    sheet = load_workbook(path, data_only=True, read_only=True).active
    rows = iter(sheet.values)
    headers = _headers(next(rows, ()))
    result: dict[str, dict[str, str]] = {}
    for values in rows:
        row = _row(headers, values)
        account = str(row.get("account") or "").strip()
        if account:
            result[account] = row
    return result


def _load_gl(path: Path) -> list[dict[str, str]]:
    with path.open("r", encoding="utf-8-sig", newline="") as stream:
        reader = csv.reader(stream)
        headers = _headers(next(reader, ()))
        return [_row(headers, values) for values in reader]


def _cash_position(path: Path | None, root: Path) -> dict[str, Any]:
    if path is None:
        return {"value": None, "complete": False, "summary": "No CFO cash-position extract was found.", "evidence": {"files": []}}
    workbook = load_workbook(path, data_only=True, read_only=True)
    sheet = next((item for item in workbook.worksheets if _normal(item.title) == "cashposition"), None)
    if sheet is None:
        return {"value": None, "complete": False, "summary": "The CFO workbook has no Cash_Position sheet.", "evidence": {"files": [_relative(path, root)]}}
    rows = iter(sheet.values)
    headers = _headers(next(rows, ()))
    entries = [_row(headers, values) for values in rows]
    dated: dict[date, list[dict[str, Any]]] = defaultdict(list)
    accounts_seen: set[str] = set()
    for row in entries:
        current = _date_value(row.get("date"))
        account = str(row.get("account") or "").strip()
        if account:
            accounts_seen.add(account)
        if current is not None:
            dated[current].append(row)
    if not dated:
        return {"value": None, "complete": False, "summary": "The Cash_Position sheet contains no dated balances.", "evidence": {"files": [_relative(path, root)]}}
    latest = max(dated)
    latest_rows = dated[latest]
    reported: list[tuple[str, Decimal]] = []
    missing: list[str] = []
    for row in latest_rows:
        account = str(row.get("account") or "Unnamed account").strip()
        value = _decimal(row.get("balance_sar"))
        if value is None:
            missing.append(account)
        else:
            reported.append((account, value))
    missing.extend(sorted(accounts_seen - {str(row.get("account") or "").strip() for row in latest_rows if row.get("account")}))
    amount = sum((value for _account, value in reported), Decimal()) if reported else None
    complete = bool(amount is not None and not missing)
    coverage = f"{len(reported)} reported balance(s)"
    if missing:
        coverage += "; missing " + ", ".join(sorted(set(missing)))
    return {
        "value": _number(amount) if amount is not None else None,
        "complete": complete,
        "summary": f"Latest cash position on {latest.isoformat()}: {_number(amount) if amount is not None else 'no reported'} SAR from {coverage}.",
        "evidence": {
            "files": [_relative(path, root)],
            "file": _relative(path, root),
            "sha256": _sha256(path),
            "sheet": sheet.title,
            "as_of": latest.isoformat(),
            "reported_accounts": [{"account": account, "balance_sar": _number(value)} for account, value in reported],
            "missing_accounts": sorted(set(missing)),
            "complete": complete,
        },
    }


def _accounts_matching(balances: Mapping[str, Decimal], accounts: Mapping[str, Mapping[str, str]], predicate) -> list[str]:
    return [account for account in sorted(balances, key=_account_sort_key) if predicate(account, accounts.get(account, {}))]


def _is_revenue(account: str, row: Mapping[str, str]) -> bool:
    return str(row.get("type") or "").lower() == "revenue" or 4000 <= _account_number(account) < 5000


def _is_cogs(account: str, _row: Mapping[str, str]) -> bool:
    return 5000 <= _account_number(account) < 6000


def _is_ebitda_exclusion(account: str, row: Mapping[str, str]) -> bool:
    text = " ".join(str(row.get(key) or "") for key in ("account_description", "type")).lower()
    return _account_number(account) in {6500, 6510, 6620} or any(term in text for term in ("depreciation", "amortization", "interest"))


def _is_operating_cost(account: str, row: Mapping[str, str]) -> bool:
    number = _account_number(account)
    return 6000 <= number < 8000 and not _is_ebitda_exclusion(account, row)


def _scope(accounts: Iterable[str], balances: Mapping[str, Decimal]) -> dict[str, Any]:
    return {"accounts": list(accounts), "net_debit_minus_credit_sar": _number(sum((balances[item] for item in accounts), Decimal()))}


def _contributors(
    scoped_accounts: Iterable[str],
    balances: Mapping[str, Decimal],
    account_master: Mapping[str, Mapping[str, str]],
    *,
    credit_nature: bool = False,
) -> list[dict[str, Any]]:
    """Return reproducible account-level contributors for executive explanation."""
    rows: list[tuple[str, Decimal]] = []
    for account in scoped_accounts:
        value = -balances[account] if credit_nature else balances[account]
        rows.append((account, value))
    total = sum((value for _account, value in rows), Decimal())
    result: list[dict[str, Any]] = []
    for account, value in sorted(rows, key=lambda item: abs(item[1]), reverse=True):
        description = str(account_master.get(account, {}).get("account_description") or "").strip()
        result.append(
            {
                "account": account,
                "label": description or f"Account {account}",
                "value_sar": _number(value),
                "share_pct": float((value / total * 100).quantize(Decimal("0.1"))) if total else None,
            }
        )
    return result


def _headers(values: Iterable[Any]) -> dict[str, int]:
    return {_normal(value): index for index, value in enumerate(values) if value is not None}


def _row(headers: Mapping[str, int], values: Iterable[Any]) -> dict[str, Any]:
    items = tuple(values)
    aliases = {
        "account": ("account",), "account_description": ("accountdescription",), "type": ("type",),
        "date": ("date",), "debit": ("debit",), "credit": ("credit",), "balance_sar": ("balancesar",),
    }
    result: dict[str, Any] = {}
    for target, choices in aliases.items():
        index = next((headers[name] for name in choices if name in headers), None)
        result[target] = items[index] if index is not None and index < len(items) else None
    return result


def _normal(value: Any) -> str:
    return "".join(character for character in str(value or "").lower() if character.isalnum())


def _decimal(value: Any) -> Decimal | None:
    if value is None or isinstance(value, bool) or str(value).strip() == "":
        return None
    try:
        return Decimal(str(value).replace(",", ""))
    except (InvalidOperation, ValueError):
        return None


def _date_value(value: Any) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    for pattern in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(str(value), pattern).date()
        except (TypeError, ValueError):
            continue
    return None


def _period_label(dates: list[date]) -> str:
    if not dates:
        return "selected period"
    start, end = min(dates), max(dates)
    if start.year == end.year and start.month == 1 and end.month == 6:
        return f"H1 {start.year}"
    return f"{start.isoformat()} to {end.isoformat()}"


def _number(value: Decimal | None) -> str | None:
    return None if value is None else format(value.quantize(Decimal("0.01")), "f")


def _sar_delta(value: Decimal) -> str:
    """Use a stable executive display while keeping the underlying trend exact."""
    sign = "+" if value > 0 else "-" if value < 0 else ""
    absolute = abs(value)
    if absolute >= Decimal("1000000"):
        amount = f"{(absolute / Decimal('1000000')):.1f}M"
    elif absolute >= Decimal("1000"):
        amount = f"{(absolute / Decimal('1000')):.1f}K"
    else:
        amount = f"{absolute.quantize(Decimal('1')):,.0f}"
    return f"{sign}SAR {amount}"


def _plan_period_fraction(period: str | None) -> Decimal | None:
    """How much of an annual plan applies to the run's reporting period.

    An annual budget is a full-year figure; the actuals here are H1. The plan
    must be scaled to the same window before the two can be compared, and the
    fraction is derived from the period label rather than assumed -- an H2 or
    full-year run must not silently borrow the H1 halving. Only period shapes
    the engine can align are accepted; anything else returns None and the plan
    stays unavailable rather than being compared across mismatched windows.
    """
    text = " ".join(str(period or "").strip().casefold().split())
    if not text:
        return None
    if "h1" in text or "h2" in text:
        return Decimal("0.5")
    if "q1" in text or "q2" in text or "q3" in text or "q4" in text:
        return Decimal("0.25")
    if "fy" in text or "full year" in text or "annual" in text:
        return Decimal("1")
    return None


def _plan_year_from_period(period: str | None) -> str | None:
    import re

    match = re.search(r"(20\d{2})", str(period or ""))
    return match.group(1) if match else None


def _reconciliation_plan(root: Path, period: str | None) -> dict[str, Any] | None:
    """Derive the like-for-like revenue plan from the governed reconciliation file.

    The dataset carries a division-to-group reconciliation stating the Central
    Region division's annual net revenue by year, including the forward budget
    (e.g. 2026F). The ERP actuals cover exactly that division, so the division
    forecast -- not the whole-BU number -- is the aligned comparator. This reads
    it, scales the matching year's forecast to the reporting window, and returns
    the plan with its own evidence.

    Every guard fails closed: no file, no forecast column, no row that names the
    division, no year match, or an unscalable period each yields None, and the
    dashboard keeps saying the comparator is unavailable rather than inventing
    one. Nothing here is derived from the actuals.
    """
    fraction = _plan_period_fraction(period)
    year = _plan_year_from_period(period)
    if fraction is None or year is None:
        return None
    path = _first_matching(root, "reconciliation", ".xlsx")
    if path is None:
        return None
    try:
        sheet = load_workbook(path, data_only=True, read_only=True).active
        rows = list(sheet.values)
    except Exception:
        return None
    if not rows:
        return None

    header = [str(c or "").strip() for c in rows[0]]
    # The forecast column is the one whose label carries the run's year and an
    # 'F' (forecast) marker -- "2026F". Actual columns ("2026A") are not a plan.
    forecast_col = None
    for idx, label in enumerate(header):
        cell = label.casefold().replace(" ", "")
        if year in cell and cell.endswith("f"):
            forecast_col = idx
            break
    if forecast_col is None:
        return None

    # The division line is the one the ERP data actually covers.
    division_annual: Decimal | None = None
    division_label = ""
    for values in rows[1:]:
        label = str((values[0] if values else "") or "")
        low = label.casefold()
        if "division" in low and "net revenue" in low and "this erp" in low:
            raw = values[forecast_col] if forecast_col < len(values) else None
            division_annual = _decimal_or_none(raw)
            division_label = label.strip()
            break
    if division_annual is None or division_annual <= 0:
        return None

    # Reconciliation states figures in SAR millions; the actuals are in SAR.
    annual_sar = (division_annual * Decimal("1000000")).quantize(Decimal("0.01"))
    period_plan = (annual_sar * fraction).quantize(Decimal("0.01"))
    return {
        "revenue_plan": _number(period_plan),
        "basis": {
            "source_file": _relative(path, root),
            "sha256": _sha256(path),
            "forecast_column": header[forecast_col],
            "division_line": division_label,
            "annual_plan_sar": _number(annual_sar),
            "period_fraction": str(fraction),
            "derivation": (
                f"{header[forecast_col]} division net revenue {_number(annual_sar)} SAR "
                f"× {fraction} of the year = {_number(period_plan)} SAR for {period}."
            ),
        },
    }


def _decimal_or_none(value: Any) -> Decimal | None:
    if value in (None, ""):
        return None
    try:
        return Decimal(str(value))
    except (InvalidOperation, TypeError, ValueError):
        return None


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        for block in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def _relative(path: Path, root: Path) -> str:
    try:
        return str(path.relative_to(root))
    except ValueError:
        return path.name


def _account_number(account: str) -> int:
    try:
        return int(account)
    except (TypeError, ValueError):
        return -1


def _account_sort_key(account: str) -> tuple[int, str]:
    return (_account_number(account), str(account))
