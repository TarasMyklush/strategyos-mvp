"""Source-bounded receivables aging with applied receipts and explicit cutoff."""
from collections import defaultdict
from decimal import Decimal, InvalidOperation
import hashlib
import json
import re
from pathlib import Path
import pandas as pd
from openpyxl import load_workbook

BUCKETS = ('Current', '1–30 days', '31–60 days', '61–90 days', '90+ days')


def _amount(value):
    try:
        result = Decimal(str(value))
        if not result.is_finite(): raise ValueError('Nonfinite money')
        return result
    except (InvalidOperation, TypeError):
        raise ValueError('Invalid monetary input')


def _date(value):
    result = pd.to_datetime(value, errors='coerce')
    if pd.isna(result): raise ValueError('Missing invoice, receipt or due date')
    return result.date()


def reconcile_aging(invoices, customers, receipts, *, as_of):
    """No settlement flag substitutes for an authoritative applied receipt."""
    segments = {}
    for row in customers:
        identity = str(row['Customer_ID']); segment = str(row.get('Segment') or 'Unassigned')
        if identity in segments and segments[identity] != segment: raise ValueError('Conflicting customer segment')
        segments[identity] = segment
    applications = defaultdict(Decimal); seen_receipts = {}; seen_invoices = {}
    for row in receipts:
        identity = str(row['Receipt_ID'])
        if identity in seen_receipts:
            if row != seen_receipts[identity]: raise ValueError('Conflicting receipt identity')
            continue
        seen_receipts[identity] = row
        if _date(row['Receipt_Date']) <= as_of:
            applications[str(row['Applied_Invoice'])] += _amount(row['Amount_SAR'])
    totals = defaultdict(lambda: {key: Decimal(0) for key in BUCKETS})
    excess = Decimal(0); count = 0; input_rows = []
    for row in invoices:
        identity = str(row['Invoice_ID'])
        if identity in seen_invoices:
            if row != seen_invoices[identity]: raise ValueError('Conflicting invoice identity')
            continue
        seen_invoices[identity] = row
        if _date(row['Invoice_Date']) > as_of: continue
        amount = _amount(row['Amount_SAR'])
        if amount < 0: raise ValueError('Credit notes need explicit application; negative invoices cannot be netted silently')
        balance = amount - applications[identity]
        if balance < 0:
            excess -= balance
            continue
        if balance == 0: continue
        days = (as_of - _date(row['Due_Date'])).days
        bucket = BUCKETS[0 if days <= 0 else 1 if days <= 30 else 2 if days <= 60 else 3 if days <= 90 else 4]
        segment = segments.get(str(row['Customer_ID']), 'Unassigned')
        totals[segment][bucket] += balance; count += 1
        input_rows.append({'invoice_id': identity, 'segment': segment, 'days_overdue': max(0,days), 'open_sar': str(balance)})
    rows = [{'segment':segment, **{key:str(value) for key,value in values.items()}, 'total_sar':str(sum(values.values()))}
            for segment,values in sorted(totals.items())]
    return {'as_of':as_of.isoformat(), 'rows':rows, 'open_invoice_count':count,
            'total_open_sar':str(sum((sum(values.values()) for values in totals.values()),Decimal(0))),
            'excess_applied_receipts_sar':str(excess), 'input_rows':input_rows}


def _verified_workbook(bundle, source):
    root = bundle.evidence.dataset_root.resolve(); path = (root/source).resolve()
    entry = bundle.evidence.manifest.get(source) or {}
    if not path.is_relative_to(root) or not path.is_file() or hashlib.sha256(path.read_bytes()).hexdigest() != entry.get('sha256'):
        raise ValueError('A required workbook is missing or changed')
    book = load_workbook(path, read_only=True, data_only=True)
    try:
        for sheet in book:
            iterator = sheet.iter_rows(values_only=True); headers = next(iterator, ())
            rows = []
            for index, values in enumerate(iterator,2):
                if index > 10000: raise ValueError('Workbook exceeds the bounded row contract')
                if any(value is not None for value in values): rows.append(dict(zip(headers,values)))
            yield sheet.title, rows
    finally: book.close()


def answer(question, bundle):
    if getattr(bundle,'evidence',None) is None: return None
    root=bundle.evidence.dataset_root
    manifest_path=root/'release-source-manifest.json'
    period=(bundle.run_metadata or {}).get('reporting_period_key')
    if not period and manifest_path.is_file(): period=json.loads(manifest_path.read_text()).get('period')
    if not re.fullmatch(r'\d{4}-\d{2}',str(period or '')):
        return {'matched':False,'available':False,'answer':'Receivables aging requires an explicit reporting cutoff and applied receipts.','basis':'The source reporting period is not supplied.','citations':[]}
    cutoff=pd.Period(period,freq='M').end_time.date()
    requested=re.search(r'\bas of (\d{4}-\d{2}-\d{2})\b',question,re.I)
    if requested:
        cutoff=_date(requested[1])
        if cutoff>pd.Period(period,freq='M').end_time.date():
            return {'matched':False,'available':False,'answer':'The requested aging date is later than the supplied source period.','basis':'Newer receipt and invoice coverage is required.','citations':[]}
    receipt_tables=[]; citations=[]
    try:
        for source in bundle.evidence.manifest:
            if not source.lower().endswith('.xlsx') or 'receipt' not in Path(source).name.lower(): continue
            for sheet, rows in _verified_workbook(bundle,source):
                if rows and {'Receipt_ID','Receipt_Date','Applied_Invoice','Amount_SAR'} <= rows[0].keys():
                    receipt_tables.extend(rows)
                    citations.append({'source_path':source,'source_hash':bundle.evidence.manifest[source]['sha256'],'locator':f'{sheet}!Excel rows 2-{len(rows)+1}','excerpt':'Applied receipts through the stated cutoff; later receipts excluded.'})
        if not receipt_tables: raise ValueError('Applied receipt coverage is not supplied')
        for role in ('ar_ledger','customer_master'):
            source=(bundle.data_contracts.get(role) or {}).get('relative_path')
            if not source: raise ValueError('Invoice or customer source contract is absent')
            sheets=list(_verified_workbook(bundle,source))
            if not sheets: raise ValueError('Invoice or customer source is empty')
            sheet,rows=sheets[0]
            citations.append({'source_path':source,'source_hash':bundle.evidence.manifest[source]['sha256'],'locator':f'{sheet}!Excel rows 2-{len(rows)+1}','excerpt':'Source invoice balances/due dates and customer segment mapping.'})
        result=reconcile_aging(bundle.ar.to_dict('records'),bundle.customers.to_dict('records'),receipt_tables,as_of=cutoff)
    except (ValueError,KeyError) as exc:
        return {'matched':False,'available':False,'answer':'Receivables aging could not be reconciled: '+str(exc)+'.','basis':'Conflicting, missing or changed source inputs block the calculation.','citations':[]}
    lines=[]
    for row in result['rows']:
        amounts='; '.join(f"{key}: SAR {Decimal(row[key]):,.2f}" for key in BUCKETS)
        lines.append(f"{row['segment']} — {amounts}; total SAR {Decimal(row['total_sar']):,.2f}.")
    ranked=sorted(result['rows'],key=lambda row:Decimal(row['90+ days']),reverse=True)
    risk=(f"The largest 90+ day exposure is in {ranked[0]['segment']} at SAR {Decimal(ranked[0]['90+ days']):,.2f}. " if ranked and Decimal(ranked[0]['90+ days'])>0 else 'No open balance is over 90 days past due. ')
    return {'matched':True,'available':True,'answer':f"As of {cutoff.isoformat()}, open receivables total SAR {Decimal(result['total_open_sar']):,.2f} across {result['open_invoice_count']} invoices.\n"+'\n'.join(lines)+'\n'+risk+'This is a dated concentration snapshot; it does not establish a worsening trend without comparable earlier snapshots.',
            'basis':'Open invoice amount less signed applied receipts through cutoff; age measured from due date. Zero balances excluded; overpayments disclosed separately; customer segments joined by Customer_ID.',
            'citations':citations,'value':result,'unit':'SAR'}
